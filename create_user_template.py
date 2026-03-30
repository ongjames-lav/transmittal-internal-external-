"""
Script to create Excel import template for users
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create a template Excel file
wb = Workbook()
ws = wb.active
ws.title = "Users"

# Define styles
header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Define columns
columns = [
    'Username',
    'Email',
    'First Name',
    'Last Name',
    'Password',
    'Contact',
    'Company',
    'Department',
    'Assigned Location',
    'Address',
    'Role',
    'Status',
]

# Write headers
for col_num, column_title in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = column_title
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Add sample rows
sample_data = [
    [
        'john.doe',
        'john@cdc.com',
        'John',
        'Doe',
        'SecurePass123',
        '09123456789',
        'CDC',
        'MIS',
        'Araneta',
        '123 Street, Araneta',
        'user',
        'pending',
    ],
    [
        'jane.smith',
        'jane@cdc.com',
        'Jane',
        'Smith',
        'SecurePass456',
        '09987654321',
        'CDC',
        'HRAD',
        'Head Office',
        '456 Ave, Manila',
        'custodian',
        'approved',
    ],
]

for row_offset, sample_data_row in enumerate(sample_data, start=2):
    for col_num, cell_value in enumerate(sample_data_row, 1):
        cell = ws.cell(row=row_offset, column=col_num)
        cell.value = cell_value

# Set column widths
widths = [15, 20, 15, 15, 16, 15, 12, 15, 18, 25, 12, 12]
for col_num, width in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + col_num)].width = width

# Save the file
template_path = 'users_import_template.xlsx'
wb.save(template_path)
print(f"✓ User import template created: {template_path}")
