"""
Admin configuration for transmittals and locations.
"""
from django.contrib import admin
from django.utils.html import format_html, mark_safe
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import path
from django.template.response import TemplateResponse
from django.contrib import messages
from .models import (
    Location, Transmittal,
    ExternalLocation, ExternalTransmittal, ExternalTransmittalAttachment, ExternalTransmittalAuditTrail
)
from accounts.models import User, Department
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io


@admin.register(Location)
class LocationAdmin(admin.ModelAdmin):
    """
    Admin interface for Location management.
    """
    list_display = [
        'name',
        'prefix',
        'custodian',
        'custodian_email_display',
        'is_active',
        'updated_at',
    ]
    
    list_filter = ['is_active']
    
    search_fields = ['name', 'prefix', 'custodian__username', 'custodian__email']
    
    list_editable = ['is_active']
    
    fieldsets = (
        ('Location Information', {
            'fields': ('name', 'prefix', 'address', 'is_active')
        }),
        ('Custodian Assignment', {
            'fields': ('custodian', 'custodian_email'),
            'description': 'Assign a custodian user to handle transmittals at this location.'
        }),
    )
    
    def custodian_email_display(self, obj):
        """Display custodian email with fallback."""
        email = obj.get_custodian_email()
        if email:
            return email
        return mark_safe('<span style="color: #999;">Not set</span>')
    custodian_email_display.short_description = 'Custodian Email'


@admin.register(Transmittal)
class TransmittalAdmin(admin.ModelAdmin):
    """
    Admin interface for Transmittal management.
    """
    change_list_template = 'admin/transmittals/transmittal/change_list.html'
    
    list_display = [
        'reference_number',
        'sender',
        'recipient_name',
        'origin_display',
        'destination_display',
        'status_badge',
        'sent_at',
    ]
    
    list_filter = [
        'status',
        'origin_location',
        'destination_location',
        'sent_at',
    ]
    
    search_fields = [
        'reference_number',
        'sender__username',
        'recipient_name',
        'recipient_email',
        'description',
    ]
    
    readonly_fields = [
        'reference_number',
        'sent_at',
        'arrived_at',
        'received_at',
        'cancelled_at',
        'arrived_by',
        'received_by',
        'date_created',
        'time_created',
        'device_information',
        'ip_address',
        'browser_of_sender',
    ]
    
    actions = ['export_to_excel']
    
    fieldsets = (
        ('Reference', {
            'fields': ('reference_number',)
        }),
        ('Sender Information', {
            'fields': ('sender', 'sender_department', 'origin_location')
        }),
        ('Recipient Information', {
            'fields': ('recipient_name', 'recipient_email', 'recipient_department', 'destination_location')
        }),
        ('Content', {
            'fields': ('description', 'remarks')
        }),
        ('Status', {
            'fields': ('status', 'is_resolved')
        }),
        ('Status Timeline', {
            'fields': ('sent_at', 'arrived_at', 'arrived_by', 'received_at', 'received_by', 'cancelled_at'),
            'classes': ('collapse',)
        }),
        ('Soft Delete', {
            'fields': ('sender_deleted', 'recipient_deleted', 'sender_purged', 'recipient_purged'),
            'classes': ('collapse',)
        }),
        ('Sender Environment', {
            'fields': ('device_information', 'ip_address', 'browser_of_sender'),
            'description': 'Device and browser information captured when the transmittal was sent',
            'classes': ('collapse',)
        }),
    )
    
    ordering = ['-sent_at']
    list_per_page = 25
    
    def get_urls(self):
        """Add custom URLs for import"""
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), 
                 name='transmittal_import_excel'),
        ]
        return custom_urls + urls
    
    def import_excel_view(self, request):
        """Handle Excel file upload and import"""
        if request.method == 'POST':
            if 'excel_file' not in request.FILES:
                messages.error(request, "❌ No file selected. Please choose an Excel file.")
                return HttpResponseRedirect(request.path)
            
            excel_file = request.FILES['excel_file']
            
            # Validate file extension
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "❌ Invalid file format. Only .xlsx and .xls files are supported.")
                return HttpResponseRedirect(request.path)
            
            # Validate file size (max 5MB)
            if excel_file.size > 5 * 1024 * 1024:
                messages.error(request, "❌ File size cannot exceed 5MB.")
                return HttpResponseRedirect(request.path)
            
            try:
                # Read Excel file
                excel_data = io.BytesIO(excel_file.read())
                workbook = load_workbook(excel_data)
                worksheet = workbook.active
                
                imported = 0
                errors = []
                warnings = []
                
                # Get all rows
                rows = list(worksheet.iter_rows(min_row=2, values_only=True))
                
                if not rows or all(cell is None for row in rows for cell in row):
                    messages.error(request, "❌ Excel file is empty. Please add data rows.")
                    return HttpResponseRedirect(request.path)
                
                # Process each row
                for row_num, row in enumerate(rows, start=2):
                    try:
                        # Skip empty rows
                        if all(cell is None for cell in row):
                            continue
                        
                        # Auto-detect format: Check if column 6 has a Department value
                        # If exported format has extra "Recipient Department" column, adjust indices
                        has_department_column = False
                        if len(row) > 6 and row[5] and len(row) > 7:
                            # Check if column 5 looks like a department (not a location)
                            col5_val = str(row[5]).strip() if row[5] else ""
                            if col5_val and col5_val not in ['ARANETA PLANT CALOOCAN', 'BAHAY PARE WAREHOUSE 17', 
                                                               'MAIN OFFICE MALABON', 'MEYC WAREHOUSE BAGBAGUIN', 'PANTOC WAREHOUSE',
                                                               'ARA', 'BPM', 'HO', 'MEY', 'PAN']:
                                # This looks like it's in the exported format with extra Department column
                                has_department_column = True
                        
                        # Map columns based on detected format
                        ref_number = row[0]
                        sender_email = row[1] if len(row) > 1 else None
                        recipient_name = row[2] if len(row) > 2 else None
                        recipient_email = row[3] if len(row) > 3 else None
                        recipient_company = row[4] if len(row) > 4 else None
                        
                        if has_department_column:
                            # Exported format: has extra Department column
                            origin_location_name = row[6] if len(row) > 6 else None
                            destination_location_name = row[7] if len(row) > 7 else None
                            status_raw = row[8] if len(row) > 8 else None
                            description = row[9] if len(row) > 9 else None
                        else:
                            # Simple import format
                            origin_location_name = row[5] if len(row) > 5 else None
                            destination_location_name = row[6] if len(row) > 6 else None
                            description = row[7] if len(row) > 7 else None
                            status_raw = row[8] if len(row) > 8 else None
                        
                        status = status_raw
                        
                        # Validation: Reference Number (Required)
                        if not ref_number or str(ref_number).strip() == '':
                            errors.append(f"Row {row_num}: ❌ Reference Number is required")
                            continue
                        
                        ref_number = str(ref_number).strip()
                        
                        # Check if reference number already exists
                        if Transmittal.objects.filter(reference_number=ref_number).exists():
                            errors.append(f"Row {row_num}: ❌ Reference number '{ref_number}' already exists in database")
                            continue
                        
                        # Validation: Sender Email (Required)
                        if not sender_email or str(sender_email).strip() == '':
                            errors.append(f"Row {row_num}: ❌ Sender Email is required")
                            continue
                        
                        sender_email = str(sender_email).strip()
                        
                        # Get sender user - try email first, then username
                        try:
                            sender = User.objects.get(email=sender_email)
                        except User.DoesNotExist:
                            # If email not found, try username
                            try:
                                sender = User.objects.get(username=sender_email)
                            except User.DoesNotExist:
                                errors.append(f"Row {row_num}: ❌ Sender '{sender_email}' (email or username) not found in system")
                                continue
                        
                        # Validation: Recipient Name (Required)
                        if not recipient_name or str(recipient_name).strip() == '':
                            errors.append(f"Row {row_num}: ❌ Recipient Name is required")
                            continue
                        
                        recipient_name = str(recipient_name).strip()
                        
                        # Validation: Recipient Email (Required)
                        if not recipient_email or str(recipient_email).strip() == '':
                            errors.append(f"Row {row_num}: ❌ Recipient Email is required")
                            continue
                        
                        recipient_email = str(recipient_email).strip()
                        
                        # Get origin location (Required)
                        if not origin_location_name or str(origin_location_name).strip() == '':
                            errors.append(f"Row {row_num}: ❌ Origin Location is required")
                            continue
                        
                        origin_location_name = str(origin_location_name).strip()
                        # Try to find location by name first, then by prefix (ONLY these 5 locations)
                        origin_location = None
                        try:
                            origin_location = Location.objects.get(name__iexact=origin_location_name)
                        except Location.DoesNotExist:
                            try:
                                origin_location = Location.objects.get(prefix__iexact=origin_location_name)
                            except Location.DoesNotExist:
                                available_locs = ', '.join([f"{loc.name} ({loc.prefix})" for loc in Location.objects.all()])
                                errors.append(f"Row {row_num}: ❌ Origin Location '{origin_location_name}' not found. Use: {available_locs}")
                                continue
                        
                        # Get destination location (Required)
                        if not destination_location_name or str(destination_location_name).strip() == '':
                            errors.append(f"Row {row_num}: ❌ Destination Location is required")
                            continue
                        
                        destination_location_name = str(destination_location_name).strip()
                        # Try to find location by name first, then by prefix (ONLY these 5 locations)
                        destination_location = None
                        try:
                            destination_location = Location.objects.get(name__iexact=destination_location_name)
                        except Location.DoesNotExist:
                            try:
                                destination_location = Location.objects.get(prefix__iexact=destination_location_name)
                            except Location.DoesNotExist:
                                available_locs = ', '.join([f"{loc.name} ({loc.prefix})" for loc in Location.objects.all()])
                                errors.append(f"Row {row_num}: ❌ Destination Location '{destination_location_name}' not found. Use: {available_locs}")
                                continue
                        
                        # Validate status (Required)
                        if not status or str(status).strip() == '':
                            errors.append(f"Row {row_num}: ❌ Status is required")
                            continue
                        
                        status = str(status).strip().lower()
                        
                        # Normalize status: convert spaces to underscores
                        # "In Transit" -> "in transit" -> "in_transit"
                        status = status.replace(' ', '_')
                        
                        valid_statuses = ['in_transit', 'arrived', 'received', 'cancelled']
                        if status not in valid_statuses:
                            errors.append(f"Row {row_num}: ❌ Status '{status}' is invalid. Valid options: {', '.join(valid_statuses)}")
                            continue
                        
                        # Create transmittal
                        transmittal = Transmittal.objects.create(
                            reference_number=ref_number,
                            sender=sender,
                            recipient_name=recipient_name,
                            recipient_email=recipient_email,
                            recipient_company=str(recipient_company).strip() if recipient_company else None,
                            origin_location=origin_location,
                            destination_location=destination_location,
                            description=str(description).strip() if description else '',
                            status=status,
                        )
                        
                        imported += 1
                        
                    except Exception as e:
                        errors.append(f"Row {row_num}: ❌ Unexpected error: {str(e)}")
                
                # Show results
                if imported > 0:
                    messages.success(
                        request, 
                        f"✓ Successfully imported {imported} transmittal(s)"
                    )
                
                if warnings:
                    for warning in warnings:
                        messages.warning(request, warning)
                
                if errors:
                    for error in errors[:10]:  # Show first 10 errors
                        messages.error(request, error)
                    if len(errors) > 10:
                        messages.error(request, f"... and {len(errors) - 10} more errors")
                
                return HttpResponseRedirect(request.path)
            
            except Exception as e:
                messages.error(request, f"❌ Error reading Excel file: {str(e)}")
                return HttpResponseRedirect(request.path)
        
        # Display import form
        context = {
            'title': 'Import Transmittals from Excel',
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
            'has_add_permission': self.has_add_permission(request),
            'available_locations': Location.objects.all(),
            'available_users': User.objects.all(),
        }
        return TemplateResponse(request, 'admin/transmittals/import_excel.html', context)
    
    def origin_display(self, obj):
        """Display origin location with prefix."""
        if obj.origin_location:
            return f"{obj.origin_location.name} ({obj.origin_location.prefix})"
        return '-'
    origin_display.short_description = 'From'
    
    def destination_display(self, obj):
        """Display destination location with prefix."""
        if obj.destination_location:
            return f"{obj.destination_location.name} ({obj.destination_location.prefix})"
        return '-'
    destination_display.short_description = 'To'
    
    def status_badge(self, obj):
        """Display colored status badge."""
        colors = {
            'in_transit': '#FFC107',   # Yellow
            'arrived': '#17A2B8',      # Blue
            'received': '#28A745',     # Green
            'cancelled': '#DC3545',    # Red
        }
        color = colors.get(obj.status, '#6C757D')
        
        return format_html(
            '<span style="background-color: {}; color: white; padding: 5px 10px; '
            'border-radius: 3px; font-weight: bold;">{}</span>',
            color,
            obj.get_status_display()
        )
    status_badge.short_description = 'Status'
    
    def export_to_excel(self, request, queryset):
        """
        Export selected transmittals to Excel file.
        """
        from django.utils import timezone
        
        # Create workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Transmittals"
        
        # Define styles
        header_fill = PatternFill(start_color="00703c", end_color="00703c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define columns
        columns = [
            'Reference Number',
            'Sender',
            'Recipient Name',
            'Recipient Email',
            'Recipient Company',
            'Recipient Department',
            'Origin Location',
            'Destination Location',
            'Status',
            'Description',
            'Remarks',
            'Sent At',
            'Arrived At',
            'Received At',
            'Cancelled At',
            'Device Information',
            'IP Address',
            'Browser of Sender',
        ]
        
        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data rows
        for row_num, transmittal in enumerate(queryset, 2):
            # Convert timestamps to local timezone (USE_TZ=False, so no conversion needed)
            sent_at_local = transmittal.sent_at if transmittal.sent_at else None
            arrived_at_local = transmittal.arrived_at if transmittal.arrived_at else None
            received_at_local = transmittal.received_at if transmittal.received_at else None
            cancelled_at_local = transmittal.cancelled_at if transmittal.cancelled_at else None
            
            data = [
                transmittal.reference_number,
                str(transmittal.sender),
                transmittal.recipient_name,
                transmittal.recipient_email,
                transmittal.recipient_company or '-',
                transmittal.recipient_department or '-',
                transmittal.origin_location.name if transmittal.origin_location else '-',
                transmittal.destination_location.name if transmittal.destination_location else '-',
                transmittal.get_status_display(),
                transmittal.description or '-',
                transmittal.remarks or '-',
                sent_at_local.strftime('%Y-%m-%d %H:%M') if sent_at_local else '-',
                arrived_at_local.strftime('%Y-%m-%d %H:%M') if arrived_at_local else '-',
                received_at_local.strftime('%Y-%m-%d %H:%M') if received_at_local else '-',
                cancelled_at_local.strftime('%Y-%m-%d %H:%M') if cancelled_at_local else '-',
                transmittal.device_information or '-',
                transmittal.ip_address or '-',
                transmittal.browser_of_sender or '-',
            ]
            
            for col_num, cell_value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 30
        ws.column_dimensions['K'].width = 30
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 18
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 18
        ws.column_dimensions['P'].width = 25  # Device Information
        ws.column_dimensions['Q'].width = 18  # IP Address
        ws.column_dimensions['R'].width = 25  # Browser of Sender
        
        # Set row heights
        ws.row_dimensions[1].height = 25
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = None  # Auto height
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Transmittals_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    
    export_to_excel.short_description = "📥 Export selected to Excel"


# ============================================================================
# EXTERNAL TRANSMITTAL ADMIN
# ============================================================================

class ExternalTransmittalAuditTrailInline(admin.TabularInline):
    """Inline display of audit trail for external transmittals."""
    model = ExternalTransmittalAuditTrail
    readonly_fields = ['action', 'performed_by', 'performed_by_email', 'timestamp', 'notes']
    extra = 0
    can_delete = False


class ExternalTransmittalAttachmentInline(admin.TabularInline):
    """Inline display of attachments for external transmittals."""
    model = ExternalTransmittalAttachment
    readonly_fields = ['file', 'attachment_type', 'uploaded_at', 'uploaded_by_email']
    extra = 0
    can_delete = False


@admin.register(ExternalLocation)
class ExternalLocationAdmin(admin.ModelAdmin):
    """Admin interface for External Location management."""
    
    list_display = [
        'name',
        'company_name',
        'email',
        'contact_person',
        'is_active',
        'updated_at',
    ]
    
    list_filter = ['is_active']
    
    search_fields = ['name', 'company_name', 'email', 'contact_person']
    
    list_editable = ['is_active']
    
    fieldsets = (
        ('Location Information', {
            'fields': ('name', 'email', 'is_active')
        }),
        ('Company Details', {
            'fields': ('company_name', 'company_address')
        }),
        ('Contact Information', {
            'fields': ('contact_person',)
        }),
    )
    
    readonly_fields = ['created_at', 'updated_at']


@admin.register(ExternalTransmittal)
class ExternalTransmittalAdmin(admin.ModelAdmin):
    """Admin interface for External Transmittal management."""
    
    list_display = [
        'reference_number',
        'main_type_display',
        'sender_email_display',
        'recipient_email',
        'status_display',
        'sub_type_display',
        'date_return',
        'date_deadline',
        'created_at',
    ]
    
    list_filter = [
        'main_type',
        'status',
        'sub_type',
        'received_status',
        'date_return',
        'date_deadline',
        'created_at',
    ]
    
    search_fields = [
        'reference_number',
        'sender_email',
        'sender_name',
        'recipient_email',
        'description',
    ]
    
    readonly_fields = [
        'reference_number',
        'created_at',
        'in_transit_at',
        'received_at',
        'closed_at',
        'last_notification_date',
        'sender_address_display',
    ]
    
    fieldsets = (
        ('Reference & Timestamps', {
            'fields': ('reference_number', 'created_at', 'in_transit_at', 'received_at', 'closed_at')
        }),
        ('Sender Information', {
            'fields': ('sender_email', 'sender_name', 'sender_company', 'sender_address_display')
        }),
        ('Recipient Information', {
            'fields': ('recipient_email', 'recipient_name', 'recipient_company_name', 'recipient_company_address')
        }),
        ('Transmittal Details', {
            'fields': ('main_type', 'sub_type', 'description', 'remarks', 'attachment')
        }),
        ('Deadline & Status', {
            'fields': ('date_return', 'date_deadline', 'status', 'received_status', 'last_notification_date')
        }),
    )
    
    inlines = [ExternalTransmittalAttachmentInline, ExternalTransmittalAuditTrailInline]
    
    def main_type_display(self, obj):
        """Display main type with color."""
        colors = {
            'for_keep': '#28A745',
            'for_return': '#FFC107',
        }
        color = colors.get(obj.main_type, '#007BFF')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color,
            obj.get_main_type_display()
        )
    main_type_display.short_description = 'Type'
    
    def sender_email_display(self, obj):
        """Display sender email (truncate if long)."""
        email = obj.sender_email
        if len(email) > 30:
            return f"{email[:30]}..."
        return email
    sender_email_display.short_description = 'Sender Email'
    
    def sender_address_display(self, obj):
        """Display sender's address from their profile."""
        try:
            user = User.objects.get(email=obj.sender_email)
            address = user.profile.address if hasattr(user, 'profile') and user.profile.address else '—'
            return address
        except User.DoesNotExist:
            return mark_safe('<span style="color: #999;">User not found</span>')
        except Exception:
            return mark_safe('<span style="color: #999;">Address not available</span>')
    sender_address_display.short_description = 'Address of the sender'
    
    def status_display(self, obj):
        """Display status with color badge."""
        colors = {
            'in_transit': '#17A2B8',
            'received': '#28A745',
            'open': '#FFC107',
            'closed': '#6C757D',
        }
        color = colors.get(obj.status, '#007BFF')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color,
            obj.get_status_display()
        )
    status_display.short_description = 'Status'
    
    def sub_type_display(self, obj):
        """Display sub type if set."""
        if obj.sub_type:
            return obj.get_sub_type_display()
        return mark_safe('<span style="color: #999;">Not set</span>')
    sub_type_display.short_description = 'Sub Type'
    
    actions = ['mark_as_closed', 'update_status_dropdown']
    
    def mark_as_closed(self, request, queryset):
        """Admin action to mark selected transmittals as closed."""
        updated = queryset.update(status='closed', received_status='closed')
        self.message_user(request, f'{updated} transmittals marked as closed.')
    mark_as_closed.short_description = 'Mark selected as Closed'
    
    def update_status_dropdown(self, request, queryset):
        """Admin action to show status update form."""
        # For simplicity, just display a message
        self.message_user(request, 'Use the admin edit form to update individual transmittal statuses.')
    update_status_dropdown.short_description = 'Update Status (use edit form)'


@admin.register(ExternalTransmittalAttachment)
class ExternalTransmittalAttachmentAdmin(admin.ModelAdmin):
    """Admin interface for External Transmittal Attachments."""
    
    list_display = [
        'transmittal',
        'attachment_type',
        'uploaded_by_email',
        'uploaded_at',
    ]
    
    list_filter = [
        'attachment_type',
        'uploaded_at',
    ]
    
    search_fields = [
        'transmittal__reference_number',
        'uploaded_by_email',
    ]
    
    readonly_fields = ['uploaded_at']
    
    fieldsets = (
        ('Attachment Information', {
            'fields': ('transmittal', 'file', 'attachment_type')
        }),
        ('Upload Details', {
            'fields': ('uploaded_by_email', 'uploaded_at')
        }),
    )


@admin.register(ExternalTransmittalAuditTrail)
class ExternalTransmittalAuditTrailAdmin(admin.ModelAdmin):
    """Admin interface for External Transmittal Audit Trail."""
    
    list_display = [
        'transmittal',
        'action_display',
        'performed_by_display',
        'timestamp',
    ]
    
    list_filter = [
        'action',
        'timestamp',
    ]
    
    search_fields = [
        'transmittal__reference_number',
        'performed_by__username',
        'performed_by_email',
    ]
    
    readonly_fields = ['transmittal', 'action', 'performed_by', 'performed_by_email', 'timestamp', 'notes']
    
    fieldsets = (
        ('Audit Information', {
            'fields': ('transmittal', 'action', 'timestamp')
        }),
        ('Performer', {
            'fields': ('performed_by', 'performed_by_email')
        }),
        ('Details', {
            'fields': ('notes', 'required_attachment_proof')
        }),
    )
    
    can_delete = False  # Prevent deletion of audit trail
    
    def action_display(self, obj):
        """Display action with styling."""
        colors = {
            'created': '#007BFF',
            'mark_received': '#28A745',
            'full_return': '#17A2B8',
            'partial_return': '#FFC107',
            'paid_sample': '#6F42C1',
            'convert_to_keep': '#28A745',
            'closed': '#6C757D',
            'admin_override': '#DC3545',
        }
        color = colors.get(obj.action, '#999')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; border-radius: 3px;">{}</span>',
            color,
            obj.get_action_display()
        )
    action_display.short_description = 'Action'
    
    def performed_by_display(self, obj):
        """Display who performed the action."""
        if obj.performed_by:
            return f"{obj.performed_by.get_full_name() or obj.performed_by.username}"
        elif obj.performed_by_email:
            return f"{obj.performed_by_email} (External)"
        return 'System'
    performed_by_display.short_description = 'Performed By'
