"""
Views for External Transmittal System.

Handles:
- Creating external transmittals (For Keep and For Return types)
- Listing inbox and sent transmittals
- Viewing transmittal details
- Resolving transmittals (mark received, full return, partial return, etc.)
- Admin status overrides
- Audit trail tracking
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.db.models import Q
from django.utils import timezone
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

from .models import (
    ExternalTransmittal, ExternalLocation, ExternalTransmittalAttachment,
    ExternalTransmittalAuditTrail
)
from .forms import ExternalTransmittalForm, ExternalTransmittalUpdateForm
from .email_utils import (
    send_external_transmittal_created_email,
    send_external_transmittal_resolution_email,
    send_external_transmittal_cancelled_email
)
from datetime import date


# ============================================================================
# EXTERNAL TRANSMITTAL DASHBOARD
# ============================================================================

@login_required(login_url='accounts:login')
@never_cache
def external_transmittal_dashboard(request):
    """
    External transmittal system dashboard.
    
    Shows statistics and quick actions for external transmittals.
    Accessible to all users (logged in or not, but stats only if logged in).
    """
    context = {
        'total_transmittals': 0,
        'for_keep_count': 0,
        'for_return_count': 0,
        'overdue_count': 0,
    }
    
    if request.user.is_authenticated:
        # Get statistics for logged-in user - strictly only their own transmittals (sender_id based)
        base_qs = ExternalTransmittal.objects.filter(
            sender_id=request.user
        )
        
        in_transit_count = base_qs.filter(status='in_transit').count()
        received_keep_count = base_qs.filter(status__in=['received', 'closed'], main_type='for_keep').count()
        received_return_count = base_qs.filter(status__in=['received', 'closed'], main_type='for_return').count()
        partial_return_count = base_qs.filter(sub_type='partial').count()
        full_return_count = base_qs.filter(sub_type='full').count()
        paid_sample_count = base_qs.filter(sub_type='for_sample').count()
        converted_keep_count = base_qs.filter(sub_type='for_keep').count()
        
        # Action required (for return but overdue, or in transit a long time, let's just use overdue logic and open received items)
        action_required = base_qs.filter(
            status__in=['in_transit', 'received'],
            received_status__in=['open', None]
        ).order_by('date_deadline')[:5]

        context.update({
            'in_transit_count': in_transit_count,
            'received_keep_count': received_keep_count,
            'received_return_count': received_return_count,
            'partial_return_count': partial_return_count,
            'full_return_count': full_return_count,
            'paid_sample_count': paid_sample_count,
            'converted_keep_count': converted_keep_count,
            'action_required': action_required,
        })
        
        # Get profile for header info
        from accounts.models import Profile
        try:
            profile = Profile.objects.select_related('department', 'assigned_location').get(user=request.user)
            context['current_department'] = profile.department
            context['current_department_name'] = profile.department.name if profile.department else 'No department assigned'
        except Profile.DoesNotExist:
            context['current_department'] = None
            context['current_department_name'] = 'No department assigned'

        # Get recent sent external transmittals (sender_id-based)
        recent_sent = ExternalTransmittal.objects.filter(
            sender_id=request.user
        ).order_by('-created_at')[:5]
        
        context.update({
            'recent_sent': recent_sent,
        })
    
    return render(request, 'transmittals/external_dashboard.html', context)


# ============================================================================
# EXTERNAL TRANSMITTAL CREATION
# ============================================================================

@login_required(login_url='accounts:login')
@never_cache
def external_transmittal_create(request):
    """
    Create a new external transmittal.
    
    Accessible without login since sender_email can be external.
    Two types supported:
    - For Keep: Permanent transfer (no return deadline needed)
    - For Return: Temporary transfer (requires deadline)
    """
    
    if request.method == 'POST':
        form = ExternalTransmittalForm(request.POST, request.FILES)
        if form.is_valid():
            # Create transmittal
            transmittal = form.save(commit=False)
            
            # Auto-fill sender information from logged-in user
            if request.user.is_authenticated:
                transmittal.sender_email = request.user.email
                transmittal.sender_name = request.user.get_full_name() or request.user.username
                transmittal.sender_id = request.user  # Set the sender_id ForeignKey
                if hasattr(request.user, 'profile'):
                    transmittal.sender_company = request.user.profile.company
            else:
                # If not logged in, sender_email comes from form
                transmittal.sender_email = transmittal.sender_email or 'external@example.com'
                transmittal.sender_name = transmittal.sender_name or 'External User'
                # sender_id remains null for external users
            

            # For Return: Always starts In-Transit; received_status set to Open
            transmittal.status = 'in_transit'
            if transmittal.main_type == 'for_return':
                transmittal.received_status = 'open'
                # Sync date_deadline with date_return since we only have one date field now
                transmittal.date_deadline = transmittal.date_return
            else:
                transmittal.received_status = None
            
            transmittal.save()
            
            # Save all uploaded attachments to ExternalTransmittalAttachment table
            uploaded_files = request.FILES.getlist('attachment')
            for uploaded_file in uploaded_files:
                ExternalTransmittalAttachment.objects.create(
                    transmittal=transmittal,
                    file=uploaded_file,
                    attachment_type='initial_upload',
                    uploaded_by_email=transmittal.sender_email
                )
            
            # Create audit trail entry
            ExternalTransmittalAuditTrail.objects.create(
                transmittal=transmittal,
                action='created',
                performed_by_email=transmittal.sender_email,
                notes=f'External transmittal created ({transmittal.main_type.upper()}) with {len(uploaded_files)} file(s)'
            )
            
            # Send notification email
            try:
                send_external_transmittal_created_email(transmittal)
            except Exception as e:
                print(f"[WARNING] Email send failed: {e}")
            
            # Redirect without success message - page auto-refreshes
            return redirect('transmittals:external_detail', pk=transmittal.pk)
    else:
        form = ExternalTransmittalForm()
    
    context = {
        'form': form,
        'page_title': 'Create External Transmittal'
    }
    return render(request, 'transmittals/external/create.html', context)


# ============================================================================
# EXTERNAL TRANSMITTAL INBOX / SENT
# ============================================================================

@login_required(login_url='accounts:login')
@never_cache
def external_transmittal_inbox(request):
    """
    List external transmittals sent by the authenticated user.

    Filters by sender_id (ID-based, not email-based).
    
    Supports hierarchical filtering via query params:
      ?status=in_transit
      ?status=received
      ?status=received&received_status=closed
      ?status=received&received_status=open
      ?status=received&received_status=open&sub_type=partial
      ?status=received&received_status=open&sub_type=for_keep
      ?status=received&received_status=open&sub_type=for_sample
    """
    # Get the authenticated user
    user = request.user

    # Base queryset: transmittals where user is the sender_id
    base_qs = ExternalTransmittal.objects.filter(
        sender_id=user
    )

    # --- URL filter params ---
    status_filter          = request.GET.get('status', '')
    received_status_filter = request.GET.get('received_status', '')
    sub_type_filter        = request.GET.get('sub_type', '')
    main_type_filter       = request.GET.get('main_type', '')

    # --- Determine active sidebar section ---
    if status_filter == 'in_transit':
        active_section = 'in_transit'
    elif status_filter == 'cancelled':
        active_section = 'cancelled'
    elif sub_type_filter == 'partial':
        active_section = 'partial'
    elif sub_type_filter == 'for_keep':
        active_section = 'for_keep_sub'
    elif sub_type_filter == 'for_sample':
        active_section = 'paid_sample'
    elif sub_type_filter == 'full':
        active_section = 'full_return'
    elif status_filter == 'received' and received_status_filter == 'closed':
        active_section = 'closed'
    elif status_filter == 'received' and received_status_filter == 'open':
        active_section = 'open'
    else:
        active_section = ''

    # --- Apply filters to queryset ---
    transmittals = base_qs.order_by('-created_at')
    
    # Handle inclusive received/closed filtering
    if status_filter == 'received':
        # If explicitly filtering for received, include closed items too
        transmittals = transmittals.filter(Q(status='received') | Q(status='closed'))
    elif status_filter:
        transmittals = transmittals.filter(status=status_filter)
        
    if received_status_filter:
        transmittals = transmittals.filter(received_status=received_status_filter)
    if sub_type_filter:
        transmittals = transmittals.filter(sub_type=sub_type_filter)
    if main_type_filter:
        transmittals = transmittals.filter(main_type=main_type_filter)

    # --- Sidebar counts ---
    count_all       = base_qs.count()
    count_intransit = base_qs.filter(status='in_transit').count()
    count_cancelled = base_qs.filter(status='cancelled').count()
    # Received main include both received-open and received-closed
    count_received  = base_qs.filter(Q(status='received') | Q(status='closed')).count()
    # Closed specifically
    count_closed    = base_qs.filter(Q(status='closed') | Q(received_status='closed')).count()
    # Open specifically
    count_open      = base_qs.filter(status='received', received_status='open').count()
    
    # Sub-types (ignore open/closed status for these specific views)
    count_partial   = base_qs.filter(sub_type='partial').count()
    count_keep      = base_qs.filter(sub_type='for_keep').count()
    count_sample    = base_qs.filter(sub_type='for_sample').count()
    count_full      = base_qs.filter(sub_type='full').count()
    
    # Main-types (counts for main_type filters)
    count_for_keep_type = base_qs.filter(main_type='for_keep').count()
    count_for_return_type = base_qs.filter(main_type='for_return').count()
    
    # For Keep with Closed status
    count_for_keep_closed = base_qs.filter(status='received', main_type='for_keep', received_status='closed').count()

    context = {
        'transmittals': transmittals,
        'status_filter': status_filter,
        'received_status_filter': received_status_filter,
        'sub_type_filter': sub_type_filter,
        'main_type_filter': main_type_filter,
        'active_section': active_section,
        'count_all': count_all,
        'count_intransit': count_intransit,
        'count_cancelled': count_cancelled,
        'count_received': count_received,
        'count_closed': count_closed,
        'count_open': count_open,
        'count_partial': count_partial,
        'count_keep': count_keep,
        'count_sample': count_sample,
        'count_full': count_full,
        'count_for_keep_type': count_for_keep_type,
        'count_for_return_type': count_for_return_type,
        'count_for_keep_closed': count_for_keep_closed,
        'page_title': 'External Transmittal Inbox',
    }
    return render(request, 'transmittals/external/inbox.html', context)


@login_required(login_url='accounts:login')
@never_cache
@login_required(login_url='accounts:login')
@never_cache
def external_transmittal_sent(request):
    """
    List external transmittals sent by the authenticated user.
    
    Shows transmittals where authenticated user's ID matches sender_id.
    This is ID-based, not email-based, for consistency with normal transmittals.
    """
    user = request.user
    
    # Get all external transmittals for this user (SENT only) - ID-based filtering
    transmittals = ExternalTransmittal.objects.filter(sender_id=user)
    
    transmittals = transmittals.order_by('-created_at')
    
    # Filter by status if provided
    status_filter = request.GET.get('status', '')
    if status_filter:
        transmittals = transmittals.filter(status=status_filter)
    
    # Filter by main_type if provided
    main_type_filter = request.GET.get('main_type', '')
    if main_type_filter:
        transmittals = transmittals.filter(main_type=main_type_filter)
    
    context = {
        'transmittals': transmittals,
        'status_filter': status_filter,
        'main_type_filter': main_type_filter,
        'page_title': 'External Transmittal Sent'
    }
    return render(request, 'transmittals/external/sent.html', context)


# ============================================================================
# EXTERNAL TRANSMITTAL DETAIL
# ============================================================================

@login_required(login_url='accounts:login')
@never_cache
def external_transmittal_detail(request, pk):
    """
    View full details of an external transmittal.
    
    Shows:
    - Transmittal information
    - Attachments
    - Audit trail history
    - Action buttons for resolution (if open)
    """
    from django.contrib.auth.models import User
    
    transmittal = get_object_or_404(ExternalTransmittal, pk=pk)
    
    # Security: Strictly only sender can view (ID-based, not email-based)
    is_sender = transmittal.sender_id == request.user
    
    if not is_sender:
        messages.error(request, "You do not have permission to view this transmittal.")
        return redirect('transmittals:external_dashboard')
    
    # Get audit trail
    audit_trail = transmittal.audit_trail.all()
    
    # Get attachments
    attachments = transmittal.attachments.all()
    
    # Get sender user from sender_id if available, otherwise from email
    sender_user = transmittal.sender_id if transmittal.sender_id else None
    if not sender_user:
        try:
            sender_user = User.objects.get(email=transmittal.sender_email)
        except User.DoesNotExist:
            sender_user = None
    
    context = {
        'transmittal': transmittal,
        'audit_trail': audit_trail,
        'attachments': attachments,
        'sender_user': sender_user,
        'page_title': f'External Transmittal {transmittal.reference_number}'
    }
    return render(request, 'transmittals/external/detail.html', context)


# ============================================================================
# EXTERNAL TRANSMITTAL RESOLUTION ACTIONS
# ============================================================================

@never_cache
@require_http_methods(['GET', 'POST'])
def external_transmittal_mark_received(request, pk):
    """
    Mark an external transmittal as received (requires proof).
    
    Applicable to both For Keep and For Return transmittals while In Transit.
    Requires proof attachment (delivery confirmation/RGA).
    Transitions status from IN_TRANSIT to RECEIVED.
    For Return: stays OPEN. For Keep: marked as CLOSED.
    """
    transmittal = get_object_or_404(ExternalTransmittal, pk=pk)
    
    # Security: Only the sender can perform this update (sender_id based)
    if not request.user.is_authenticated or transmittal.sender_id != request.user:
        messages.error(request, 'Only the original sender can mark this transmittal as received.')
        return redirect('transmittals:external_detail', pk=pk)
    
    # Verify it can be marked received
    if not transmittal.can_mark_received():
        messages.error(request, 'This action is only available for transmittals in transit.')
        return redirect('transmittals:external_detail', pk=pk)
    
    if request.method == 'POST':
        form = ExternalTransmittalUpdateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Create attachments with proof - handle multiple files
                uploaded_files = request.FILES.getlist('attachment')
                for uploaded_file in uploaded_files:
                    ExternalTransmittalAttachment.objects.create(
                        transmittal=transmittal,
                        file=uploaded_file,
                        attachment_type='Proof of Delivery',
                        uploaded_by_email=request.user.email
                    )
                
                # Update transmittal status
                transmittal.status = 'received'
                transmittal.received_at = timezone.now()
                
                if transmittal.is_for_keep():
                    transmittal.received_status = 'closed'
                else:
                    transmittal.received_status = 'open'
                
                transmittal.save()
                
                # Create audit trail
                ExternalTransmittalAuditTrail.objects.create(
                    transmittal=transmittal,
                    action='mark_received',
                    performed_by_email=request.user.email,
                    notes=form.cleaned_data.get('notes', ''),
                )
                
                # Send notification email
                try:
                    send_external_transmittal_resolution_email(
                        transmittal=transmittal,
                        action_type='mark_received',
                        notes=form.cleaned_data.get('notes', '')
                    )
                except Exception as e:
                    print(f"[WARNING] Email send failed: {e}")
                
                # Redirect without success message - page auto-refreshes
                return redirect('transmittals:external_detail', pk=pk)
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        form = ExternalTransmittalUpdateForm()
        form.fields['action'].initial = 'mark_received'
        # Set file input label specific to this action
        form.fields['attachment'].label = 'Proof of Delivery / Received RGA (REQUIRED)'
    
    context = {
        'form': form,
        'transmittal': transmittal,
        'action_title': 'Mark as Received',
        'page_title': f'Mark {transmittal.reference_number} as Received'
    }
    return render(request, 'transmittals/external/action.html', context)


@never_cache
@require_http_methods(['GET', 'POST'])
def external_transmittal_full_return(request, pk):
    """
    Mark a For Return external transmittal as fully returned (requires proof).
    """
    transmittal = get_object_or_404(ExternalTransmittal, pk=pk)
    
    # Security: Only the sender can perform this update (sender_id based)
    if not request.user.is_authenticated or transmittal.sender_id != request.user:
        messages.error(request, 'Only the original sender can record a return.')
        return redirect('transmittals:external_detail', pk=pk)

    # Verify transition
    if not transmittal.can_transition_to_full_return():
        messages.error(request, 'This action is only available for open transmittals marked as received.')
        return redirect('transmittals:external_detail', pk=pk)
    
    if request.method == 'POST':
        form = ExternalTransmittalUpdateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Create attachments with proof - handle multiple files
                uploaded_files = request.FILES.getlist('attachment')
                for uploaded_file in uploaded_files:
                    ExternalTransmittalAttachment.objects.create(
                        transmittal=transmittal,
                        file=uploaded_file,
                        attachment_type='Proof of Full Return',
                        uploaded_by_email=request.user.email
                    )
                
                # Update transmittal
                transmittal.sub_type = 'full'
                transmittal.received_status = 'closed'
                transmittal.status = 'closed'
                transmittal.closed_at = timezone.now()
                transmittal.save()
                
                # Create audit trail
                ExternalTransmittalAuditTrail.objects.create(
                    transmittal=transmittal,
                    action='full_return',
                    performed_by_email=request.user.email,
                    notes=form.cleaned_data.get('notes', '')
                )
                try:
                    send_external_transmittal_resolution_email(
                        transmittal=transmittal,
                        action_type='full_return',
                        notes=form.cleaned_data.get('notes', '')
                    )
                except Exception as e:
                    print(f"[WARNING] Email send failed: {e}")
                
                # Redirect without success message - page auto-refreshes
                return redirect('transmittals:external_detail', pk=pk)
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        form = ExternalTransmittalUpdateForm()
        form.fields['action'].initial = 'full_return'
        form.fields['attachment'].label = 'Proof of Full Return (REQUIRED)'
    
    context = {
        'form': form,
        'transmittal': transmittal,
        'action_title': 'Full Return',
        'page_title': f'Record Full Return for {transmittal.reference_number}'
    }
    return render(request, 'transmittals/external/action.html', context)


@never_cache
@require_http_methods(['GET', 'POST'])
def external_transmittal_partial_return(request, pk):
    """
    Mark a For Return external transmittal as partially returned.
    """
    transmittal = get_object_or_404(ExternalTransmittal, pk=pk)
    
    # Security: Only the sender can perform this update (sender_id based)
    if not request.user.is_authenticated or transmittal.sender_id != request.user:
        messages.error(request, 'Only the original sender can record a partial return.')
        return redirect('transmittals:external_detail', pk=pk)

    # Verify transition
    if not transmittal.can_transition_to_partial():
        messages.error(request, 'This action is only available for open transmittals marked as received.')
        return redirect('transmittals:external_detail', pk=pk)
    
    if request.method == 'POST':
        form = ExternalTransmittalUpdateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Create attachments with proof - handle multiple files
                uploaded_files = request.FILES.getlist('attachment')
                for uploaded_file in uploaded_files:
                    ExternalTransmittalAttachment.objects.create(
                        transmittal=transmittal,
                        file=uploaded_file,
                        attachment_type='Proof of Partial Return',
                        uploaded_by_email=request.user.email
                    )
                
                # Update transmittal - remains open for further resolution
                transmittal.sub_type = 'partial'
                transmittal.received_status = 'open'
                transmittal.status = 'received'
                transmittal.save()
                
                # Create audit trail
                ExternalTransmittalAuditTrail.objects.create(
                    transmittal=transmittal,
                    action='partial_return',
                    performed_by_email=request.user.email,
                    notes=form.cleaned_data.get('notes', '')
                )
                
                # Send notification email
                try:
                    send_external_transmittal_resolution_email(
                        transmittal=transmittal,
                        action_type='partial_return',
                        notes=form.cleaned_data.get('notes', '')
                    )
                except Exception as e:
                    print(f"[WARNING] Email send failed: {e}")
                
                # Redirect without success message - page auto-refreshes
                return redirect('transmittals:external_detail', pk=pk)
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        form = ExternalTransmittalUpdateForm()
        form.fields['action'].initial = 'partial_return'
        form.fields['attachment'].label = 'Proof of Partial Return (REQUIRED)'
    
    context = {
        'form': form,
        'transmittal': transmittal,
        'action_title': 'Partial Return',
        'page_title': f'Record Partial Return for {transmittal.reference_number}'
    }
    return render(request, 'transmittals/external/action.html', context)


@never_cache
@require_http_methods(['GET', 'POST'])
def external_transmittal_paid_sample(request, pk):
    """
    Convert a transmittal to Paid Sample (receiver keeps items and pays).
    """
    transmittal = get_object_or_404(ExternalTransmittal, pk=pk)
    
    # Security: Only the sender can perform this update (sender_id based)
    if not request.user.is_authenticated or transmittal.sender_id != request.user:
        messages.error(request, 'Only the original sender can change this to Paid Sample.')
        return redirect('transmittals:external_detail', pk=pk)

    # Verify transition
    if not transmittal.can_transition_to_paid_sample():
        messages.error(request, 'This action is only available for open transmittals marked as received.')
        return redirect('transmittals:external_detail', pk=pk)
    
    if request.method == 'POST':
        form = ExternalTransmittalUpdateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Create attachments with RGA proof - handle multiple files
                uploaded_files = request.FILES.getlist('attachment')
                for uploaded_file in uploaded_files:
                    ExternalTransmittalAttachment.objects.create(
                        transmittal=transmittal,
                        file=uploaded_file,
                        attachment_type='RGA / Deduction Proof',
                        uploaded_by_email=request.user.email
                    )
                
                # Update transmittal
                transmittal.sub_type = 'for_sample'
                transmittal.received_status = 'closed'
                transmittal.status = 'closed'
                transmittal.closed_at = timezone.now()
                transmittal.save()
                
                # Create audit trail
                ExternalTransmittalAuditTrail.objects.create(
                    transmittal=transmittal,
                    action='paid_sample',
                    performed_by_email=request.user.email,
                    notes=form.cleaned_data.get('notes', '')
                )
                
                # Send notification email
                try:
                    send_external_transmittal_resolution_email(
                        transmittal=transmittal,
                        action_type='paid_sample',
                        notes=form.cleaned_data.get('notes', '')
                    )
                except Exception as e:
                    print(f"[WARNING] Email send failed: {e}")
                
                # Redirect without success message - page auto-refreshes
                return redirect('transmittals:external_detail', pk=pk)
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        form = ExternalTransmittalUpdateForm()
        form.fields['action'].initial = 'paid_sample'
        form.fields['attachment'].label = 'RGA / Deduction Proof (REQUIRED)'
    
    context = {
        'form': form,
        'transmittal': transmittal,
        'action_title': 'Paid Sample',
        'page_title': f'{transmittal.reference_number} - Paid Sample'
    }
    return render(request, 'transmittals/external/action.html', context)


@never_cache
@require_http_methods(['GET', 'POST'])
def external_transmittal_convert_to_keep(request, pk):
    """
    Convert a for return transmittal to For Keep (SubType).
    """
    transmittal = get_object_or_404(ExternalTransmittal, pk=pk)
    
    # Security: Only the sender can perform this update (sender_id based)
    if not request.user.is_authenticated or transmittal.sender_id != request.user:
        messages.error(request, 'Only the original sender can change this to For Keep.')
        return redirect('transmittals:external_detail', pk=pk)

    # Verify transition
    if not transmittal.can_transition_to_for_keep_subtype():
        messages.error(request, 'This action is only available for open transmittals marked as received.')
        return redirect('transmittals:external_detail', pk=pk)
    
    if request.method == 'POST':
        form = ExternalTransmittalUpdateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Create attachments with RGA proof - handle multiple files
                uploaded_files = request.FILES.getlist('attachment')
                for uploaded_file in uploaded_files:
                    ExternalTransmittalAttachment.objects.create(
                        transmittal=transmittal,
                        file=uploaded_file,
                        attachment_type='RGA Proof',
                        uploaded_by_email=request.user.email
                    )
                
                # Update transmittal
                transmittal.sub_type = 'for_keep'
                transmittal.received_status = 'closed'
                transmittal.status = 'closed'
                transmittal.closed_at = timezone.now()
                transmittal.save()
                
                # Create audit trail
                ExternalTransmittalAuditTrail.objects.create(
                    transmittal=transmittal,
                    action='convert_to_keep',
                    performed_by_email=request.user.email,
                    notes=form.cleaned_data.get('notes', '')
                )
                
                # Send notification email
                try:
                    send_external_transmittal_resolution_email(
                        transmittal=transmittal,
                        action_type='convert_to_keep',
                        notes=form.cleaned_data.get('notes', '')
                    )
                except Exception as e:
                    print(f"[WARNING] Email send failed: {e}")
                
                # Redirect without success message - page auto-refreshes
                return redirect('transmittals:external_detail', pk=pk)
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        form = ExternalTransmittalUpdateForm()
        form.fields['action'].initial = 'convert_to_keep'
        form.fields['attachment'].label = 'RGA Proof (REQUIRED)'
    
    context = {
        'form': form,
        'transmittal': transmittal,
        'action_title': 'For Keep',
        'page_title': f'{transmittal.reference_number} - For Keep'
    }
    return render(request, 'transmittals/external/action.html', context)


# ============================================================================
# ADMIN ONLY: STATUS OVERRIDE
# ============================================================================

@staff_member_required
@require_http_methods(['POST'])
def external_transmittal_admin_override(request, pk):
    """
    Admin-only action to override transmittal status.
    
    Allows admins to manually change status via dropdown.
    Logs action to audit trail for compliance.
    """
    transmittal = get_object_or_404(ExternalTransmittal, pk=pk)
    
    new_status = request.POST.get('status')
    reason = request.POST.get('reason', '')
    
    if new_status not in ['in_transit', 'received', 'open', 'closed']:
        messages.error(request, 'Invalid status.')
        return redirect('transmittals:external_detail', pk=pk)
    
    old_status = transmittal.status
    transmittal.status = new_status
    transmittal.save()
    
    # Log admin override to audit trail
    ExternalTransmittalAuditTrail.objects.create(
        transmittal=transmittal,
        action='admin_override',
        performed_by=request.user,
        notes=f'Status changed from {old_status} to {new_status}. Reason: {reason}'
    )
    
    messages.success(request, f'Status updated to {transmittal.get_status_display()}')
    return redirect('transmittals:external_detail', pk=pk)

# ============================================================================
# EXTERNAL TRANSMITTAL CANCELLATION
# ============================================================================

@login_required(login_url='accounts:login')
@never_cache
@require_http_methods(['GET', 'POST'])
def cancel_external_transmittal(request, pk):
    """
    Cancel an external transmittal (sender only, in_transit status only).
    
    GET: Display cancellation confirmation form
    POST: Process cancellation with reason
    """
    transmittal = get_object_or_404(ExternalTransmittal, pk=pk)
    
    # Authorization: only sender can cancel
    if transmittal.sender_id != request.user:
        messages.error(request, 'You are not authorized to cancel this transmittal.')
        return redirect('transmittals:external_inbox')
    
    # Can only cancel if in transit
    if not transmittal.can_cancel():
        messages.error(request, 'This transmittal cannot be cancelled. It must be in "In Transit" status.')
        return redirect('transmittals:external_detail', pk=pk)
    
    if request.method == 'POST':
        cancellation_reason = request.POST.get('cancellation_reason', '').strip()
        
        if not cancellation_reason:
            messages.error(request, 'Please provide a cancellation reason.')
            return render(request, 'transmittals/external/cancel.html', {
                'transmittal': transmittal,
                'form_error': 'Cancellation reason is required.'
            })
        
        # Update transmittal
        transmittal.status = 'cancelled'
        transmittal.received_status = 'cancelled'
        transmittal.cancelled_at = timezone.now()
        transmittal.cancellation_reason = cancellation_reason
        transmittal.save()
        
        # Create audit trail entry
        ExternalTransmittalAuditTrail.objects.create(
            transmittal=transmittal,
            action='cancelled',
            performed_by_email=request.user.email,
            notes=f'Transmittal cancelled by sender. Reason: {cancellation_reason}'
        )
        
        # Send cancellation email to recipient
        try:
            send_external_transmittal_cancelled_email(transmittal, cancellation_reason)
        except Exception as e:
            print(f"[WARNING] Failed to send cancellation email: {e}")
        
        messages.success(request, f'Transmittal {transmittal.reference_number} has been cancelled successfully.')
        return redirect('transmittals:external_inbox')
    
    # GET: Show cancellation form
    context = {
        'transmittal': transmittal,
        'page_title': f'Cancel - {transmittal.reference_number}'
    }
    return render(request, 'transmittals/external/cancel.html', context)


# ============================================================================
# EXTERNAL TRANSMITTAL REPORT & EXPORT
# ============================================================================

def export_external_to_excel(transmittals):
    """Export external transmittals to Excel file."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "External Transmittals"

        # Define styles
        header_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Define headers
        headers = [
            'Reference Number',
            'Type',
            'Sub Type',
            'Status',
            'Sender Name',
            'Sender Company',
            'Recipient Name',
            'Recipient Email',
            'Recipient Company',
            'Recipient Address',
            'Description',
            'Remarks',
            'Created Date',
            'Received Date',
            'Closed Date',
            'Return Date',
            'Deadline Date',
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Add data
        transmittal_list = list(transmittals)
        for row_idx, t in enumerate(transmittal_list, 2):
            try:
                # Get sub type display
                sub_type_display = ''
                if t.sub_type:
                    sub_type_map = {
                        'full': 'Full Return',
                        'partial': 'Partial Return',
                        'for_sample': 'Paid Sample',
                        'for_keep': 'For Keep (Sub)',
                    }
                    sub_type_display = sub_type_map.get(t.sub_type, t.sub_type)
                
                # Get sender name (not UUID)
                sender_name = str(t.sender_name) if t.sender_name else ''
                if not sender_name and t.sender_id:
                    # Fallback to user's full name if sender_name is empty
                    try:
                        sender_name = t.sender_id.get_full_name() or t.sender_id.username
                    except:
                        sender_name = ''

                ws.cell(row=row_idx, column=1).value = str(t.reference_number) if t.reference_number else ''
                ws.cell(row=row_idx, column=2).value = 'For Keep' if t.main_type == 'for_keep' else 'For Return'
                ws.cell(row=row_idx, column=3).value = sub_type_display
                ws.cell(row=row_idx, column=4).value = str(t.get_status_display()).upper() if t.status else ''
                ws.cell(row=row_idx, column=5).value = sender_name
                ws.cell(row=row_idx, column=6).value = str(t.sender_company) if t.sender_company else ''
                ws.cell(row=row_idx, column=7).value = str(t.recipient_name) if t.recipient_name else ''
                ws.cell(row=row_idx, column=8).value = str(t.recipient_email) if t.recipient_email else ''
                ws.cell(row=row_idx, column=9).value = str(t.recipient_company_name) if t.recipient_company_name else ''
                ws.cell(row=row_idx, column=10).value = str(t.recipient_company_address) if t.recipient_company_address else ''
                ws.cell(row=row_idx, column=11).value = str(t.description) if t.description else ''
                ws.cell(row=row_idx, column=12).value = str(t.remarks) if t.remarks else ''
                ws.cell(row=row_idx, column=13).value = t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else ''
                ws.cell(row=row_idx, column=14).value = t.received_at.strftime('%Y-%m-%d %H:%M') if t.received_at else ''
                ws.cell(row=row_idx, column=15).value = t.closed_at.strftime('%Y-%m-%d %H:%M') if t.closed_at else ''
                ws.cell(row=row_idx, column=16).value = t.date_return.strftime('%Y-%m-%d') if t.date_return else ''
                ws.cell(row=row_idx, column=17).value = t.date_deadline.strftime('%Y-%m-%d') if t.date_deadline else ''

                # Apply borders and alignment
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border
                    cell.alignment = left_align
            except Exception as e:
                print(f"ERROR processing external transmittal row {row_idx}: {e}")
                import traceback
                traceback.print_exc()
                continue

        # Adjust column widths
        col_widths = [20, 12, 16, 14, 20, 20, 20, 25, 20, 30, 35, 25, 18, 18, 18, 14, 14]
        for i, width in enumerate(col_widths):
            ws.column_dimensions[chr(65 + i)].width = width

        # Create response
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_content = excel_buffer.getvalue()
        excel_buffer.seek(0)

        filename = f'External_Transmittals_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            excel_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
        response['Content-Length'] = len(excel_content)
        response['X-Content-Type-Options'] = 'nosniff'
        print(f"DEBUG: External export file created: {filename}, size: {len(excel_content)} bytes")
        return response
    except Exception as e:
        print(f"CRITICAL ERROR in external export: {e}")
        import traceback
        traceback.print_exc()
        raise


@login_required(login_url='accounts:login')
@never_cache
def external_transmittal_report(request):
    """
    Report page for external transmittals.

    Each user sees only their own transmittals (sender_id = request.user).
    Supports filtering by status, main_type, sub_type, and date range.
    Supports Excel export with ?download=excel query param.
    """
    user = request.user

    # Base queryset: only the current user's transmittals
    transmittals = ExternalTransmittal.objects.filter(
        sender_id=user
    ).order_by('-created_at')

    # Get filter parameters
    status_filter = request.GET.get('status', '')
    main_type_filter = request.GET.get('main_type', '')
    sub_type_filter = request.GET.get('sub_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Apply status filter
    if status_filter:
        if status_filter == 'received':
            transmittals = transmittals.filter(Q(status='received') | Q(status='closed'))
        else:
            transmittals = transmittals.filter(status=status_filter)

    # Apply main type filter
    if main_type_filter:
        transmittals = transmittals.filter(main_type=main_type_filter)

    # Apply sub type filter
    if sub_type_filter:
        transmittals = transmittals.filter(sub_type=sub_type_filter)

    # Apply date range filter
    if date_from:
        try:
            from datetime import datetime as dt
            date_from_obj = dt.strptime(date_from, '%Y-%m-%d')
            transmittals = transmittals.filter(created_at__gte=date_from_obj)
        except Exception:
            pass

    if date_to:
        try:
            from datetime import datetime as dt, timedelta
            date_to_obj = dt.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            transmittals = transmittals.filter(created_at__lt=date_to_obj)
        except Exception:
            pass

    # Handle Excel export download
    if request.GET.get('download') == 'excel':
        try:
            response = export_external_to_excel(transmittals)
            return response
        except Exception as e:
            print(f"ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error generating Excel file: {str(e)}")

    context = {
        'transmittals': transmittals,
        'transmittal_count': transmittals.count(),
        'status_filter': status_filter,
        'main_type_filter': main_type_filter,
        'sub_type_filter': sub_type_filter,
        'date_from': date_from,
        'date_to': date_to,
        'page_title': 'External Transmittal Reports',
    }
    return render(request, 'transmittals/external/report.html', context)


# ============================================================================
# SECURE FILE DOWNLOAD (REQUIRES LOGIN)
# ============================================================================

@login_required(login_url='accounts:login')
@never_cache
def secure_external_file_download(request, transmittal_id, attachment_id):
    """
    Secure file download endpoint requiring authentication.
    
    Validates that:
    1. User is logged in
    2. User is the sender of the transmittal
    3. Attachment belongs to the transmittal
    
    Returns the file or 403 Forbidden if unauthorized.
    """
    from django.core.files.storage import default_storage
    import os
    
    # Get transmittal and verify ownership
    transmittal = get_object_or_404(ExternalTransmittal, pk=transmittal_id)
    
    # Security: Only sender can download
    if transmittal.sender_id != request.user:
        messages.error(request, 'You do not have permission to download this file.')
        return redirect('transmittals:external_dashboard')
    
    # Get attachment and verify it belongs to transmittal
    attachment = get_object_or_404(
        ExternalTransmittalAttachment,
        pk=attachment_id,
        transmittal=transmittal
    )
    
    # Serve the file
    if attachment.file:
        file_path = attachment.file.path
        
        # Verify file exists
        if not os.path.exists(file_path):
            messages.error(request, 'File not found.')
            return redirect('transmittals:external_detail', pk=transmittal_id)
        
        # Get file info
        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path)
        
        # Serve file
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            response['Content-Length'] = file_size
            return response
    
    messages.error(request, 'File not available.')
    return redirect('transmittals:external_detail', pk=transmittal_id)
