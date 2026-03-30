from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.db.models import Q
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

from .models import Transmittal, Location, ExternalTransmittal
from .forms import TransmittalForm, StatusUpdateForm, CancelTransmittalForm, ReceiveTransmittalForm
from .email_utils import send_transmittal_email, send_status_notification, send_cancellation_email, send_cancellation_email_to_receiver
from .environment_utils import capture_sender_environment
from accounts.models import Profile


# ============================================================
# DASHBOARD
# ============================================================
@never_cache
@login_required(login_url='accounts:login')
def dashboard(request):
    """
    User dashboard showing recent transmittal status.
    """
    # Recent sent
    recent_sent = Transmittal.objects.filter(
        sender=request.user,
        sender_deleted=False
    ).order_by('-sent_at')[:5]
    
    # Recent received
    recent_received = Transmittal.objects.filter(
        recipient_id=request.user,
        recipient_deleted=False
    ).order_by('-sent_at')[:5]
    
    # Counts by status
    sent_count = Transmittal.objects.filter(sender=request.user, sender_deleted=False).count()
    received_count = Transmittal.objects.filter(recipient_id=request.user, recipient_deleted=False).count()
    
    context = {
        'recent_sent': recent_sent,
        'recent_received': recent_received,
        'sent_count': sent_count,
        'received_count': received_count,
    }
    return render(request, 'transmittals/dashboard.html', context)


# ============================================================
# CREATE TRANSMITTAL
# ============================================================
@never_cache
@login_required(login_url='accounts:login')
def create_transmittal(request):
    """
    Create new transmittal with auto-fill fields.
    
    Auto-fills: sender, sender_department, origin_location, reference_number
    User inputs: recipient info, destination_location, description, remarks
    """
    # Check if user is approved, admin, or custodian
    if not (request.user.is_staff or request.user.profile.status == 'approved' or is_custodian(request.user)):
        messages.error(request, "Your account must be approved to send transmittal reports.")
        return redirect('accounts:dashboard')
    
    # Get user's profile for auto-fill
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found. Please contact admin.")
        return redirect('accounts:dashboard')
    
    # Get user's origin location - use assigned_location first (works for all users)
    origin_location = None
    
    # Try assigned location first
    if profile.assigned_location:
        origin_location = profile.assigned_location
    else:
        # Fall back to matching profile location to Location object
        origin_location = Location.objects.filter(
            Q(name__iexact=profile.location) | Q(prefix__iexact=profile.location),
            is_active=True
        ).first()
    
    # If still no match, default to first active location
    if not origin_location:
        origin_location = Location.objects.filter(is_active=True).first()
    
    if request.method == 'POST':
        form = TransmittalForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            transmittal = form.save(commit=False)
            
            # Get selected user from profile
            selected_user_profile = form.cleaned_data.get('selected_user')
            if selected_user_profile:
                transmittal.recipient_name = selected_user_profile.get_full_name()
                transmittal.recipient_email = selected_user_profile.user.email
                transmittal.recipient_id = selected_user_profile.user
                transmittal.recipient_company = selected_user_profile.company
                transmittal.recipient_department = selected_user_profile.department
                # Set destination location from recipient's assigned location
                if selected_user_profile.assigned_location:
                    transmittal.destination_location = selected_user_profile.assigned_location
            else:
                messages.error(request, 'Please select a recipient.')
                return render(request, 'transmittals/create_transmittal.html', {'form': form, 'user': request.user})
            
            # Auto-fill sender information
            transmittal.sender = request.user
            transmittal.sender_department = profile.department
            transmittal.origin_location = origin_location
            
            # Set status: custodians send transmittals as 'arrived', others as 'in_transit'
            if is_custodian(request.user):
                transmittal.status = 'arrived'
                transmittal.arrived_at = datetime.now()
                transmittal.arrived_by = request.user
            else:
                transmittal.status = 'in_transit'
            
            # Capture sender environment details
            environment = capture_sender_environment(request)
            transmittal.device_information = environment['device_information']
            transmittal.ip_address = environment['ip_address']
            transmittal.browser_of_sender = environment['browser_of_sender']
            
            # Save to generate reference number
            transmittal.save()
            
            # Handle multiple file attachments
            attachments = request.FILES.getlist('attachments')
            if attachments:
                from .models import Attachment
                for file in attachments:
                    Attachment.objects.create(
                        transmittal=transmittal,
                        file=file
                    )
            
            # If this is a reverse transmittal, copy attachments from original
            reverse_from_id = request.POST.get('reverse_from_id')
            if reverse_from_id:
                try:
                    original_transmittal = Transmittal.objects.get(pk=reverse_from_id)
                    original_attachments = original_transmittal.attachments.all()
                    if original_attachments:
                        from .models import Attachment
                        for original_attachment in original_attachments:
                            # Copy the attachment
                            Attachment.objects.create(
                                transmittal=transmittal,
                                file=original_attachment.file
                            )
                except Transmittal.DoesNotExist:
                    pass  # If original transmittal doesn't exist, just skip attachment copying
            
            
            # Send email notifications
            email_sent = send_transmittal_email(transmittal)
            
            if email_sent:
                messages.success(request, f'Transmittal {transmittal.reference_number} sent successfully!')
            else:
                messages.warning(request, f'Transmittal {transmittal.reference_number} saved but email delivery failed.')
            
            # Store transmittal ID in session for success page
            request.session['last_transmittal_id'] = transmittal.id
            return redirect('transmittals:transmittal_success')
    else:
        # Check if this is a reverse transmittal request
        reverse_from_id = request.GET.get('reverse_from')
        initial_data = {}
        selected_user_name = None
        
        if reverse_from_id:
            try:
                original_transmittal = Transmittal.objects.get(pk=reverse_from_id)
                
                # Only recipient can create reverse transmittal
                if original_transmittal.recipient_id != request.user:
                    messages.error(request, "You can only reverse transmittals sent to you.")
                    return redirect('transmittals:inbox')
                
                # Get the original sender
                sender = original_transmittal.sender
                
                # Try to get the sender's profile to set as selected_user
                try:
                    sender_profile = Profile.objects.get(user=sender)
                    initial_data['selected_user'] = sender_profile.id
                    selected_user_name = sender_profile.get_full_name()
                except Profile.DoesNotExist:
                    pass
                
                # Pre-fill data for reverse transmittal
                initial_data.update({
                    'destination_location': original_transmittal.origin_location.id if original_transmittal.origin_location else None,
                    'assigned_location_text': original_transmittal.origin_location.name if original_transmittal.origin_location else '',
                    'recipient_name': original_transmittal.sender.profile.get_full_name(),
                    'recipient_email': original_transmittal.sender.email,
                    'recipient_department': original_transmittal.sender_department,
                    'description': original_transmittal.description,
                    'remarks': original_transmittal.remarks if original_transmittal.remarks else '',
                })
                
                # Also try to get and pre-fill the sender's profile information for company and location
                sender_profile = original_transmittal.sender.profile
                if sender_profile:
                    # Add company if available
                    if hasattr(sender_profile, 'company') and sender_profile.company:
                        initial_data['recipient_company'] = sender_profile.company
                    
                    # Add assigned location if available
                    if sender_profile.assigned_location:
                        initial_data['destination_location'] = sender_profile.assigned_location.id
            except Transmittal.DoesNotExist:
                messages.error(request, "Original transmittal not found.")
                return redirect('transmittals:inbox')
        
        form = TransmittalForm(user=request.user, initial=initial_data)
    
    # Generate preview reference number
    preview_ref = Transmittal().generate_reference_number() if origin_location else "REF-XXXXXXXX-XXXX"
    if origin_location:
        # Create a temp object to generate preview
        temp = Transmittal(origin_location=origin_location)
        preview_ref = temp.generate_reference_number()
    
    # Get all active locations for dropdown
    locations = Location.objects.filter(is_active=True)
    
    # Get all approved users for recipient selection
    approved_users = Profile.objects.filter(status='approved').select_related('user', 'assigned_location')
    
    # If user is custodian, filter approved_users to only those at the same location
    if is_custodian(request.user):
        custodian_location = profile.assigned_location
        if custodian_location:
            approved_users = approved_users.filter(assigned_location=custodian_location)
    
    context = {
        'form': form,
        'user': request.user,
        'profile': profile,
        'origin_location': origin_location,
        'locations': locations,
        'approved_users': approved_users,
        'preview_ref': preview_ref,
        'today': datetime.now().strftime('%B %d, %Y'),
        'now': datetime.now().strftime('%I:%M %p'),
        'selected_user_name': selected_user_name,
    }
    return render(request, 'transmittals/create_transmittal.html', context)


@login_required(login_url='accounts:login')
def transmittal_success(request):
    """Display success message after transmittal submission with print option."""
    transmittal_id = request.session.get('last_transmittal_id')
    transmittal = None
    
    if transmittal_id:
        transmittal = Transmittal.objects.filter(id=transmittal_id).first()
        # Clear session after fetching
        del request.session['last_transmittal_id']
    
    return render(request, 'transmittals/transmittal_success.html', {
        'transmittal': transmittal
    })


# ============================================================
# TRANSMITTAL LISTS
# ============================================================
@never_cache
@login_required(login_url='accounts:login')
def inbox(request):
    """Display received transmittals (Transmittal Report Received)."""
    transmittals = Transmittal.objects.filter(
        recipient_id=request.user,
        recipient_deleted=False
    )
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter in ['arrived', 'picked', 'received', 'cancelled', 'in_transit']:
        transmittals = transmittals.filter(status=status_filter)
    else:
        # By default, show only in_transit items (inbox)
        transmittals = transmittals.filter(status='in_transit')

    # Handle sorting
    sort_by = request.GET.get('sort', '-sent_at')
    sort_direction = '↑' if sort_by.startswith('-') else '↓'
    sort_field = sort_by.lstrip('-')
    
    # Validate sort field
    valid_sorts = {
        'sender__profile__user__first_name': 'sender name',
        'description': 'description',
        'reference_number': 'reference number',
        'sent_at': 'time sent',
        '-sender__profile__user__first_name': 'sender name',
        '-description': 'description',
        '-reference_number': 'reference number',
        '-sent_at': 'time sent',
    }
    
    if sort_by not in valid_sorts:
        # Also support recipient_name for flexibility
        if sort_by in ['recipient_name', '-recipient_name']:
             # In inbox, sorting by recipient name is usually constant (me), but allowed
             pass
        else:
            sort_by = '-sent_at'
    
    transmittals = transmittals.order_by(sort_by)
    
    # Determine next sort direction
    next_sort = sort_by if sort_by.startswith('-') else f"-{sort_by}"
    
    return render(request, 'transmittals/inbox.html', {
        'transmittals': transmittals,
        'folder': 'inbox',
        'title': 'Transmittal Reports Received',
        'current_sort': sort_by,
        'sort_direction': sort_direction,
        'sort_field': sort_field,
        'status_filter': status_filter
    })


@never_cache
@login_required(login_url='accounts:login')
def sent_emails(request):
    """Display sent transmittals (Transmittal Report Sent) - all transmittals sent by user."""
    from datetime import datetime, timedelta
    
    transmittals = Transmittal.objects.filter(
        sender=request.user,
        sender_deleted=False
    )
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter in ['arrived', 'picked', 'received', 'cancelled', 'in_transit']:
        transmittals = transmittals.filter(status=status_filter)
    else:
        # By default, exclude cancelled transmittals (show only active ones)
        transmittals = transmittals.exclude(status='cancelled')
    
    # Filter by date range if provided
    date_filter = request.GET.get('date_filter')
    today = datetime.now().date()
    
    if date_filter == 'today':
        transmittals = transmittals.filter(sent_at__date=today)
    elif date_filter == 'week':
        start_date = today - timedelta(days=7)
        transmittals = transmittals.filter(sent_at__date__gte=start_date)
    elif date_filter == 'month':
        start_date = today - timedelta(days=30)
        transmittals = transmittals.filter(sent_at__date__gte=start_date)
    elif date_filter == 'custom':
        # Handle custom date range
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
                transmittals = transmittals.filter(sent_at__date__gte=start_dt)
            except ValueError:
                pass
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
                transmittals = transmittals.filter(sent_at__date__lte=end_dt)
            except ValueError:
                pass
    
    # Handle sorting
    sort_by = request.GET.get('sort', '-sent_at')
    sort_direction = '↑' if sort_by.startswith('-') else '↓'
    sort_field = sort_by.lstrip('-')
    
    # Validate sort field
    valid_sorts = {
        'recipient_name': 'recipient name',
        'description': 'description',
        'reference_number': 'reference number',
        'sent_at': 'time sent',
        '-recipient_name': 'recipient name',
        '-description': 'description',
        '-reference_number': 'reference number',
        '-sent_at': 'time sent',
    }
    
    if sort_by not in valid_sorts:
         # Check for sender name compatibility
        if sort_by in ['sender__profile__user__first_name', '-sender__profile__user__first_name']:
            pass
        else:
            sort_by = '-sent_at'
    
    transmittals = transmittals.order_by(sort_by)
    
    return render(request, 'transmittals/inbox.html', {
        'transmittals': transmittals,
        'folder': 'sent',
        'title': 'Transmittal Reports Sent',
        'current_sort': sort_by,
        'sort_direction': sort_direction,
        'sort_field': sort_field,
        'status_filter': status_filter,
        'date_filter': date_filter,
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
        'all_statuses': ['in_transit', 'arrived', 'received']
    })


# ============================================================
# TRANSMITTAL DETAIL & STATUS
# ============================================================
@never_cache
@login_required(login_url='accounts:login')
def transmittal_detail(request, pk):
    """View details of a specific transmittal with status actions."""
    transmittal = get_object_or_404(Transmittal, pk=pk)
    
    # Refresh sender profile from database to get latest info
    sender_profile = Profile.objects.select_related('user').get(user=transmittal.sender)
    sender_full_name = sender_profile.get_full_name()
    sender_department = sender_profile.department.name if sender_profile.department else 'N/A'
    
    # Security check
    is_sender = transmittal.sender == request.user
    is_recipient = transmittal.recipient_id == request.user
    
    # Check if user is custodian for destination or origin location
    is_custodian = False
    is_origin_custodian = False
    is_destination_custodian = False
    try:
        if request.user.profile.role == 'custodian':
            profile = request.user.profile
            # Check if custodian is for destination location
            if transmittal.destination_location and profile.assigned_location == transmittal.destination_location:
                is_custodian = True
                is_destination_custodian = True
            # Check if custodian is for origin location
            elif transmittal.origin_location and profile.assigned_location == transmittal.origin_location:
                is_custodian = True
                is_origin_custodian = True
    except Profile.DoesNotExist:
        pass
    
    # Allow staff to view anything
    if not (is_sender or is_recipient or is_custodian or request.user.is_staff):
        messages.error(request, "You do not have permission to view this transmittal.")
        return redirect('transmittals:inbox')
    
    # Determine available actions
    can_cancel = is_sender and transmittal.can_cancel()
    # Only destination custodian can mark arrived (not origin custodian)
    can_mark_arrived = is_destination_custodian and transmittal.status == 'in_transit'
    # Only destination custodian can mark as picked (from 'arrived' status)
    can_mark_pick = is_destination_custodian and transmittal.status == 'arrived'
    # Receiver can mark as received from both 'in_transit' and 'arrived' statuses
    # Only recipients (not custodians) can mark as received
    if is_recipient:
        can_mark_received = transmittal.status in ['in_transit', 'arrived', 'picked']
    else:
        can_mark_received = False
    
    # Receiver can reverse transmittal (create new one with swapped origin/destination)
    # They can do this at any point to send something back
    can_reverse = is_recipient
    
    # Pre-format dates to avoid template formatting issues
    # Since USE_TZ = False, we work with naive datetimes directly
    if transmittal.sent_at:
        sent_at_formatted = transmittal.sent_at.strftime('%B %d, %Y %I:%M %p')
    else:
        sent_at_formatted = ''
    
    if transmittal.received_at:
        received_at_formatted = transmittal.received_at.strftime('%B %d, %Y at %I:%M %p')
    else:
        received_at_formatted = ''
    
    if transmittal.arrived_at:
        arrived_at_formatted = transmittal.arrived_at.strftime('%B %d, %Y at %I:%M %p')
    else:
        arrived_at_formatted = ''
    
    if transmittal.picked_at:
        picked_at_formatted = transmittal.picked_at.strftime('%B %d, %Y at %I:%M %p')
    else:
        picked_at_formatted = ''
    
    context = {
        'transmittal': transmittal,
        'sender_full_name': sender_full_name,
        'sender_department': sender_department,
        'is_sender': is_sender,
        'is_recipient': is_recipient,
        'is_custodian': is_custodian,
        'can_cancel': can_cancel,
        'can_mark_arrived': can_mark_arrived,
        'can_mark_pick': can_mark_pick,
        'can_mark_received': can_mark_received,
        'can_reverse': can_reverse,
        'sent_at_formatted': sent_at_formatted,
        'received_at_formatted': received_at_formatted,
        'arrived_at_formatted': arrived_at_formatted,
        'picked_at_formatted': picked_at_formatted,
    }
    return render(request, 'transmittals/detail.html', context)


@never_cache
@login_required(login_url='accounts:login')
def mark_arrived(request, pk):
    """Custodian marks transmittal as arrived."""
    transmittal = get_object_or_404(Transmittal, pk=pk)
    
    # Check if user is a custodian with the proper location assignment
    try:
        profile = request.user.profile
        is_custodian_role = profile.role == 'custodian'
        # Check if destination location matches custodian's assigned location
        is_correct_location = (transmittal.destination_location and 
                              profile.assigned_location == transmittal.destination_location)
        is_custodian = is_custodian_role and is_correct_location
    except:
        is_custodian = False
    
    if not is_custodian:
        messages.error(request, "Only the custodian can mark transmittals as arrived.")
        # For quick actions, redirect to custodian section
        if request.GET.get('quick') == 'true':
            return redirect('transmittals:custodian_in_transit')
        return redirect('transmittals:detail', pk=pk)
    
    if transmittal.status != 'in_transit':
        messages.error(request, "This transmittal is not in transit.")
        return redirect('transmittals:detail', pk=pk)
    
    # Handle quick action from list (GET request with quick=true)
    if request.method == 'GET' and request.GET.get('quick') == 'true':
        transmittal.status = 'arrived'
        transmittal.arrived_at = datetime.now()
        transmittal.arrived_by = request.user
        transmittal.save()
        
        # Send notification
        send_status_notification(transmittal, 'arrived')
        
        messages.success(request, f"Transmittal {transmittal.reference_number} marked as Arrived.")
        return redirect('transmittals:custodian_arrived')
    
    if request.method == 'POST':
        transmittal.status = 'arrived'
        transmittal.arrived_at = datetime.now()
        transmittal.arrived_by = request.user
        transmittal.save()
        
        # Send notification
        send_status_notification(transmittal, 'arrived')
        
        messages.success(request, f"Transmittal {transmittal.reference_number} marked as Arrived.")
        return redirect('transmittals:detail', pk=pk)
    
    return render(request, 'transmittals/confirm_status.html', {
        'transmittal': transmittal,
        'action': 'arrived',
        'action_display': 'Arrived'
    })


@never_cache
@login_required(login_url='accounts:login')
def mark_pick(request, pk):
    """Custodian marks transmittal as picked up with optional remarks."""
    transmittal = get_object_or_404(Transmittal, pk=pk)
    
    # Check if user is a custodian with the proper location assignment
    try:
        profile = request.user.profile
        is_custodian_role = profile.role == 'custodian'
        # Check if destination location matches custodian's assigned location
        is_correct_location = (transmittal.destination_location and 
                              profile.assigned_location == transmittal.destination_location)
        is_custodian = is_custodian_role and is_correct_location
    except:
        is_custodian = False
    
    if not is_custodian:
        messages.error(request, "Only the custodian can mark transmittals as picked.")
        # For quick actions, redirect to custodian section
        if request.GET.get('quick') == 'true':
            return redirect('transmittals:custodian_arrived')
        return redirect('transmittals:detail', pk=pk)
    
    if transmittal.status != 'arrived':
        messages.error(request, "This transmittal is not in arrived status.")
        return redirect('transmittals:detail', pk=pk)
    
    # Handle quick action from list (GET request with quick=true)
    if request.method == 'GET' and request.GET.get('quick') == 'true':
        transmittal.status = 'picked'
        transmittal.picked_at = datetime.now()
        transmittal.save()
        
        # Send notification
        send_status_notification(transmittal, 'picked')
        
        messages.success(request, f"Transmittal {transmittal.reference_number} marked as Picked.")
        return redirect('transmittals:custodian_picked')
    
    if request.method == 'POST':
        # Get optional remarks from form
        pick_remarks = request.POST.get('pick_remarks', '').strip()
        
        transmittal.status = 'picked'
        transmittal.picked_at = datetime.now()
        if pick_remarks:
            transmittal.pick_remarks = pick_remarks
        transmittal.save()
        
        # Send notification
        send_status_notification(transmittal, 'picked')
        
        messages.success(request, f"Transmittal {transmittal.reference_number} marked as Picked.")
        return redirect('transmittals:detail', pk=pk)
    
    return render(request, 'transmittals/mark_pick.html', {
        'transmittal': transmittal,
        'action': 'picked',
        'action_display': 'Picked'
    })


@never_cache
@login_required(login_url='accounts:login')
def mark_received(request, pk):
    """Receiver/Custodian marks transmittal as received.
    
    Custodian can mark as received from arrived status
    Receiver can mark as received from both in_transit and arrived statuses
    """
    transmittal = get_object_or_404(Transmittal, pk=pk)
    
    # Check if user is sender, custodian or recipient
    is_sender = transmittal.sender == request.user
    is_custodian = False
    try:
        profile = request.user.profile
        if profile.role == 'custodian':
            is_custodian = (transmittal.destination_location and 
                          profile.assigned_location == transmittal.destination_location)
    except:
        pass
    
    is_recipient = transmittal.recipient_id == request.user
    
    if not (is_sender or is_custodian or is_recipient):
        messages.error(request, "You don't have permission to mark this transmittal as received.")
        # For quick actions, redirect to custodian section
        if request.GET.get('quick') == 'true':
            return redirect('transmittals:custodian_arrived')
        return redirect('transmittals:detail', pk=pk)
    
    # Check status based on user role
    if is_recipient:
        # Receiver can mark from 'in_transit', 'arrived', or 'picked'
        if transmittal.status not in ['in_transit', 'arrived', 'picked']:
            messages.error(request, "This transmittal cannot be marked as received.")
            return redirect('transmittals:detail', pk=pk)
    else:
        # Custodian and Sender can only mark from 'arrived'
        if transmittal.status != 'arrived':
            messages.error(request, "You can only mark as received when the status is 'Arrived'.")
            return redirect('transmittals:detail', pk=pk)
    
    # Handle quick action from list (GET request with quick=true)
    if request.method == 'GET' and request.GET.get('quick') == 'true':
        transmittal.status = 'received'
        transmittal.received_at = datetime.now()
        transmittal.received_by = request.user
        transmittal.auto_received = False  # Manual receive by user
        transmittal.is_resolved = True
        
        # Attach user's signature if they have one
        user_profile = request.user.profile
        if user_profile.digital_signature:
            transmittal.receiver_signature = user_profile.digital_signature
        
        transmittal.save()
        
        # Send notification to sender and custodian (if custodian exists)
        send_status_notification(transmittal, 'received')
        
        messages.success(request, f"Transmittal {transmittal.reference_number} marked as Received.")
        return redirect('transmittals:inbox')
    
    if request.method == 'POST':
        form = ReceiveTransmittalForm(request.POST, request.FILES)
        if form.is_valid():
            transmittal.status = 'received'
            transmittal.received_at = datetime.now()
            transmittal.received_by = request.user
            transmittal.auto_received = False  # Manual receive by user
            transmittal.is_resolved = True
            
            # Handle signature upload
            signature = form.cleaned_data.get('signature')
            if signature:
                transmittal.receiver_signature = signature
            else:
                # Use user's stored signature if available
                user_profile = request.user.profile
                if user_profile.digital_signature:
                    transmittal.receiver_signature = user_profile.digital_signature
            
            transmittal.save()
            
            # Send notification to sender and custodian (if custodian exists)
            send_status_notification(transmittal, 'received')
            
            messages.success(request, f"Transmittal {transmittal.reference_number} marked as Received.")
            return redirect('transmittals:detail', pk=pk)
    else:
        form = ReceiveTransmittalForm()
    
    return render(request, 'transmittals/confirm_status.html', {
        'transmittal': transmittal,
        'action': 'received',
        'action_display': 'Received',
        'form': form
    })


@never_cache
@login_required(login_url='accounts:login')
def cancel_transmittal(request, pk):
    """Sender cancels transmittal (only if still in transit)."""
    transmittal = get_object_or_404(Transmittal, pk=pk)
    
    # Check if user is sender
    if transmittal.sender != request.user:
        messages.error(request, "Only the sender can cancel transmittals.")
        return redirect('transmittals:detail', pk=pk)
    
    if not transmittal.can_cancel():
        messages.error(request, "This transmittal can no longer be cancelled.")
        return redirect('transmittals:detail', pk=pk)
    
    if request.method == 'POST':
        form = CancelTransmittalForm(request.POST)
        if form.is_valid():
            cancellation_reason = form.cleaned_data.get('reason', 'No reason provided')
            
            transmittal.status = 'cancelled'
            transmittal.cancelled_at = datetime.now()
            transmittal.cancellation_reason = cancellation_reason
            transmittal.save()
            
            # Send cancellation emails
            send_cancellation_email(transmittal, cancellation_reason)
            send_cancellation_email_to_receiver(transmittal, cancellation_reason)
            
            messages.success(request, f"Transmittal {transmittal.reference_number} has been cancelled and notifications have been sent.")
            return redirect(f"{reverse('transmittals:sent_emails')}?status=cancelled")
    else:
        form = CancelTransmittalForm()
    
    return render(request, 'transmittals/cancel_transmittal.html', {
        'transmittal': transmittal,
        'form': form
    })


# ============================================================
# PRINT VIEW
# ============================================================
@login_required(login_url='accounts:login')
def print_transmittal(request, pk):
    """Print-friendly view of transmittal report."""
    transmittal = get_object_or_404(Transmittal, pk=pk)
    
    # Security check
    is_sender = transmittal.sender == request.user
    is_recipient = transmittal.recipient_id == request.user
    
    # Check if user is custodian for destination or origin location
    custodian_user = False
    try:
        if request.user.profile.role == 'custodian':
            profile = request.user.profile
            # Custodian can view if destination location matches their assigned location
            if transmittal.destination_location and profile.assigned_location == transmittal.destination_location:
                custodian_user = True
            # Custodian can also view if origin location matches their assigned location
            elif transmittal.origin_location and profile.assigned_location == transmittal.origin_location:
                custodian_user = True
    except Profile.DoesNotExist:
        pass
    
    is_custodian = custodian_user
    
    if not (is_sender or is_recipient or is_custodian or request.user.is_staff):
        messages.error(request, "You do not have permission to view this transmittal.")
        return redirect('transmittals:inbox')
    
    return render(request, 'transmittals/print.html', {
        'transmittal': transmittal
    })


# ============================================================
# TRASH MANAGEMENT - DISABLED
# ============================================================

@never_cache
@login_required(login_url='accounts:login')
def delete_emails(request):
    """Handle deletion of selected transmittals (soft delete)."""
    if request.method == 'POST':
        email_ids = request.POST.getlist('email_ids')
        folder = request.POST.get('folder')
        
        if email_ids:
            if folder == 'inbox':
                transmittals = Transmittal.objects.filter(
                    id__in=email_ids, 
                    recipient_id=request.user
                )
                count = transmittals.update(
                    recipient_deleted=True,
                    recipient_deleted_at=datetime.now()
                )
                messages.success(request, f"{count} transmittal(s) moved to Trash.")
            elif folder == 'sent':
                transmittals = Transmittal.objects.filter(
                    id__in=email_ids, 
                    sender=request.user
                )
                count = transmittals.update(
                    sender_deleted=True,
                    sender_deleted_at=datetime.now()
                )
                messages.success(request, f"{count} transmittal(s) moved to Trash.")
            else:
                messages.error(request, "Invalid folder.")
        else:
            messages.warning(request, "No transmittals selected for deletion.")
            
        if folder == 'sent':
            return redirect('transmittals:sent_emails')
        return redirect('transmittals:inbox')
        
    return redirect('transmittals:inbox')


@never_cache
@login_required(login_url='accounts:login')
def restore_emails(request):
    """Restore selected emails from Trash."""
    if request.method == 'POST':
        email_ids = request.POST.getlist('email_ids')
        
        if email_ids:
            rec_count = Transmittal.objects.filter(
                id__in=email_ids,
                recipient_id=request.user,
                recipient_deleted=True
            ).update(recipient_deleted=False, recipient_deleted_at=None)
            
            send_count = Transmittal.objects.filter(
                id__in=email_ids,
                sender=request.user,
                sender_deleted=True
            ).update(sender_deleted=False, sender_deleted_at=None)
            
            total = rec_count + send_count
            messages.success(request, f"{total} transmittal(s) restored.")
            
    return redirect('transmittals:inbox')


@never_cache
@login_required(login_url='accounts:login')
def permanent_delete_emails(request):
    """Permanently delete (purge) selected emails from Trash."""
    if request.method == 'POST':
        email_ids = request.POST.getlist('email_ids')
        
        if email_ids:
            rec_count = Transmittal.objects.filter(
                id__in=email_ids,
                recipient_id=request.user,
                recipient_deleted=True
            ).update(recipient_purged=True)
            
            send_count = Transmittal.objects.filter(
                id__in=email_ids,
                sender=request.user,
                sender_deleted=True
            ).update(sender_purged=True)
            
            total = rec_count + send_count
            messages.success(request, f"{total} transmittal(s) permanently deleted.")
            
    return redirect('transmittals:inbox')


# ============================================================
# AJAX ENDPOINTS
# ============================================================
@login_required(login_url='accounts:login')
def get_location_custodian(request, location_id):
    """AJAX endpoint to get custodian info for a location."""
    try:
        location = Location.objects.get(id=location_id, is_active=True)
        custodian_name = ''
        custodian_email = ''
        
        if location.custodian:
            custodian_name = location.custodian.get_full_name() or location.custodian.username
            custodian_email = location.get_custodian_email() or ''
        
        return JsonResponse({
            'success': True,
            'custodian_name': custodian_name,
            'custodian_email': custodian_email,
            'location_name': location.name,
            'location_prefix': location.prefix,
        })
    except Location.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Location not found'
        })

@login_required(login_url='accounts:login')
def search_suggestions(request):
    """
    AJAX endpoint for search suggestions.
    Supports both internal and external transmittal systems.
    """
    query = request.GET.get('q', '')
    system = request.GET.get('system', 'internal')
    
    if len(query) < 2:
        return JsonResponse({'results': []})
        
    user = request.user
    
    if system == 'external':
        # Search ExternalTransmittal - strictly only sender_id based (no recipient_email filtering)
        base_qs = ExternalTransmittal.objects.filter(
            sender_id=user
        )
        
        # 1. Reference Number Matches
        ref_matches = base_qs.filter(
            reference_number__icontains=query
        ).values('id', 'reference_number', 'recipient_name', 'description')[:5]
        
        # 2. Recipient Name Matches
        name_matches = base_qs.filter(
            recipient_name__icontains=query
        ).exclude(
            id__in=[r['id'] for r in ref_matches]
        ).values('id', 'reference_number', 'recipient_name', 'description')[:5]
        
        # 3. Recipient Email Matches
        email_matches = base_qs.filter(
            recipient_email__icontains=query
        ).exclude(
            id__in=[r['id'] for r in ref_matches]
        ).exclude(
            id__in=[r['id'] for r in name_matches]
        ).values('id', 'reference_number', 'recipient_name', 'recipient_email', 'description')[:5]
        
        # 4. Description Matches
        desc_matches = base_qs.filter(
            description__icontains=query
        ).exclude(
            id__in=[r['id'] for r in ref_matches]
        ).exclude(
            id__in=[r['id'] for r in name_matches]
        ).exclude(
            id__in=[r['id'] for r in email_matches]
        ).values('id', 'reference_number', 'recipient_name', 'description')[:5]

        results = []
        for r in ref_matches:
            results.append({
                'id': r['id'],
                'title': r['reference_number'],
                'subtitle': f"To: {r['recipient_name'] or 'External'}",
                'type': 'Reference'
            })
        for r in name_matches:
            results.append({
                'id': r['id'],
                'title': f"To: {r['recipient_name']}",
                'subtitle': r['reference_number'],
                'type': 'Person'
            })
        for r in email_matches:
            results.append({
                'id': r['id'],
                'title': r['recipient_email'],
                'subtitle': f"{r['recipient_name'] or 'External'} - {r['reference_number']}",
                'type': 'Email'
            })
        for r in desc_matches:
            results.append({
                'id': r['id'],
                'title': r['description'][:100] + '...' if len(r['description']) > 100 else r['description'],
                'subtitle': r['reference_number'],
                'type': 'Description'
            })
            
        return JsonResponse({'results': results})

    else:
        # Default: Search Internal Transmittal
        if user.is_staff or user.is_superuser:
            base_qs = Transmittal.objects.all()
        elif is_custodian(user):
            try:
                location = user.profile.assigned_location
                base_qs = Transmittal.objects.filter(destination_location=location)
            except (Profile.DoesNotExist, AttributeError):
                base_qs = Transmittal.objects.none()
        else:
            base_qs = Transmittal.objects.filter(
                Q(sender=user) | Q(recipient_id=user)
            )
        
        base_qs = base_qs.distinct()
        
        ref_matches = base_qs.filter(reference_number__icontains=query).values('id', 'reference_number', 'recipient_name', 'description')[:5]
        name_matches = base_qs.filter(recipient_name__icontains=query).exclude(id__in=[r['id'] for r in ref_matches]).values('id', 'reference_number', 'recipient_name', 'description')[:5]
        email_matches = base_qs.filter(recipient_email__icontains=query).exclude(id__in=[r['id'] for r in ref_matches]).exclude(id__in=[r['id'] for r in name_matches]).values('id', 'reference_number', 'recipient_name', 'recipient_email', 'description')[:5]
        desc_matches = base_qs.filter(description__icontains=query).exclude(id__in=[r['id'] for r in ref_matches]).exclude(id__in=[r['id'] for r in name_matches]).exclude(id__in=[r['id'] for r in email_matches]).values('id', 'reference_number', 'recipient_name', 'description')[:5]
        
        results = []
        for r in ref_matches:
            results.append({'id': r['id'], 'title': r['reference_number'], 'subtitle': f"To: {r['recipient_name']}", 'type': 'Reference'})
        for r in name_matches:
            results.append({'id': r['id'], 'title': f"To: {r['recipient_name']}", 'subtitle': r['reference_number'], 'type': 'Person'})
        for r in email_matches:
            results.append({'id': r['id'], 'title': r['recipient_email'], 'subtitle': f"{r['recipient_name']} - {r['reference_number']}", 'type': 'Email'})
        for r in desc_matches:
            results.append({'id': r['id'], 'title': r['description'][:100] + '...' if len(r['description']) > 100 else r['description'], 'subtitle': r['reference_number'], 'type': 'Description'})
            
        return JsonResponse({'results': results})


@login_required(login_url='accounts:login')
def update_driver_remarks(request, pk):
    """AJAX endpoint to update driver remarks for a transmittal."""
    if request.method == 'POST':
        transmittal = get_object_or_404(Transmittal, pk=pk)
        
        # Check permissions - only custodian can update driver info
        is_custodian = False
        try:
            if request.user.profile.role == 'custodian':
                profile = request.user.profile
                if transmittal.origin_location and profile.assigned_location == transmittal.origin_location:
                    is_custodian = True
        except Profile.DoesNotExist:
            pass
        
        if not (is_custodian or request.user.is_staff):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to update driver information.'
            }, status=403)
        
        driver_remarks = request.POST.get('driver_remarks', '').strip()
        
        if not driver_remarks:
            return JsonResponse({
                'success': False,
                'error': 'Driver information cannot be empty.'
            })
        
        try:
            transmittal.driver_remarks = driver_remarks
            transmittal.save()
            
            # Send email notification to receiver and destination custodian
            from .email_utils import send_driver_update_email
            email_sent = send_driver_update_email(transmittal, driver_remarks, request.user)
            
            print(f"[DEBUG] Driver update email sent: {email_sent} for transmittal {transmittal.pk}")
            
            return JsonResponse({
                'success': True,
                'message': 'Driver information updated successfully.'
            })
        except Exception as e:
            print(f"[ERROR] Exception in update_driver_remarks: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method.'
    }, status=400)


@login_required(login_url='accounts:login')
def bulk_update_driver_remarks(request):
    """AJAX endpoint to update driver remarks for multiple transmittals."""
    if request.method == 'POST':
        import json
        
        driver_remarks = request.POST.get('driver_remarks', '').strip()
        transmittal_ids_str = request.POST.get('transmittal_ids', '[]')
        
        if not driver_remarks:
            return JsonResponse({
                'success': False,
                'error': 'Driver information cannot be empty.'
            })
        
        try:
            transmittal_ids = json.loads(transmittal_ids_str)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid transmittal IDs.'
            })
        
        if not transmittal_ids:
            return JsonResponse({
                'success': False,
                'error': 'No transmittals selected.'
            })
        
        # Check permissions - only custodian can update driver info for their location
        is_custodian = False
        custodian_location = None
        try:
            if request.user.profile.role == 'custodian':
                is_custodian = True
                custodian_location = request.user.profile.assigned_location
        except Profile.DoesNotExist:
            pass
        
        if not (is_custodian or request.user.is_staff):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to update driver information.'
            }, status=403)
        
        try:
            updated_count = 0
            from .email_utils import send_driver_update_email
            
            for transmittal_id in transmittal_ids:
                try:
                    transmittal = Transmittal.objects.get(pk=transmittal_id)
                    
                    # For non-admin, verify they can update this transmittal (custodian of origin location)
                    if not request.user.is_staff:
                        if not transmittal.origin_location or transmittal.origin_location != custodian_location:
                            continue
                    
                    # Update driver remarks
                    transmittal.driver_remarks = driver_remarks
                    transmittal.save()
                    
                    # Send email notification
                    send_driver_update_email(transmittal, driver_remarks, request.user)
                    updated_count += 1
                    
                    print(f"[DEBUG] Bulk driver update for transmittal {transmittal.pk}")
                except Transmittal.DoesNotExist:
                    continue
            
            if updated_count == 0:
                return JsonResponse({
                    'success': False,
                    'error': 'No transmittals were updated. Please check your permissions.'
                })
            
            return JsonResponse({
                'success': True,
                'message': f'Driver information updated for {updated_count} transmittal(s).',
                'updated_count': updated_count
            })
        except Exception as e:
            print(f"[ERROR] Exception in bulk_update_driver_remarks: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method.'
    }, status=400)


@login_required(login_url='accounts:login')
@require_http_methods(["POST"])
def bulk_pick(request):
    """AJAX endpoint to mark multiple transmittals as picked."""
    if request.method == 'POST':
        import json
        
        pick_remarks = request.POST.get('pick_remarks', '').strip()
        transmittal_ids_str = request.POST.get('transmittal_ids', '[]')
        
        try:
            transmittal_ids = json.loads(transmittal_ids_str)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid transmittal IDs.'
            })
        
        if not transmittal_ids:
            return JsonResponse({
                'success': False,
                'error': 'No transmittals selected.'
            })
        
        # Check permissions - only custodian of destination location can pick
        is_custodian = False
        custodian_location = None
        try:
            if request.user.profile.role == 'custodian':
                is_custodian = True
                custodian_location = request.user.profile.assigned_location
        except Profile.DoesNotExist:
            pass
        
        if not (is_custodian or request.user.is_staff):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to pick transmittals.'
            }, status=403)
        
        try:
            updated_count = 0
            from django.utils import timezone
            
            for transmittal_id in transmittal_ids:
                try:
                    transmittal = Transmittal.objects.get(pk=transmittal_id)
                    
                    # Verify transmittal is in arrived status
                    if transmittal.status != 'arrived':
                        continue
                    
                    # For non-admin, verify they can pick this transmittal (custodian of destination location)
                    if not request.user.is_staff:
                        if not transmittal.destination_location or transmittal.destination_location != custodian_location:
                            continue
                    
                    # Update status to picked
                    transmittal.status = 'picked'
                    transmittal.picked_at = timezone.now()
                    if pick_remarks:
                        transmittal.pick_remarks = pick_remarks
                    transmittal.save()
                    
                    print(f"[DEBUG] Bulk pick for transmittal {transmittal.pk}")
                    updated_count += 1
                    
                except Transmittal.DoesNotExist:
                    continue
            
            if updated_count == 0:
                return JsonResponse({
                    'success': False,
                    'error': 'No transmittals were picked. Please check your permissions.'
                })
            
            return JsonResponse({
                'success': True,
                'message': f'{updated_count} transmittal(s) picked successfully.',
                'updated_count': updated_count
            })
        except Exception as e:
            print(f"[ERROR] Exception in bulk_pick: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method.'
    }, status=400)


@login_required(login_url='accounts:login')
@require_http_methods(["POST"])
def bulk_arrived(request):
    """AJAX endpoint to mark multiple transmittals as arrived."""
    if request.method == 'POST':
        import json
        
        remarks = request.POST.get('remarks', '').strip()
        transmittal_ids_str = request.POST.get('transmittal_ids', '[]')
        
        try:
            transmittal_ids = json.loads(transmittal_ids_str)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid transmittal IDs.'
            })
        
        if not transmittal_ids:
            return JsonResponse({
                'success': False,
                'error': 'No transmittals selected.'
            })
        
        # Check permissions - only custodian of destination location can mark as arrived
        is_custodian = False
        custodian_location = None
        try:
            if request.user.profile.role == 'custodian':
                is_custodian = True
                custodian_location = request.user.profile.assigned_location
        except Profile.DoesNotExist:
            pass
        
        if not (is_custodian or request.user.is_staff):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to mark transmittals as arrived.'
            }, status=403)
        
        try:
            updated_count = 0
            from django.utils import timezone
            
            for transmittal_id in transmittal_ids:
                try:
                    transmittal = Transmittal.objects.get(pk=transmittal_id)
                    
                    # Verify transmittal is in in_transit status
                    if transmittal.status != 'in_transit':
                        continue
                    
                    # For non-admin, verify they can mark this transmittal (custodian of destination location)
                    if not request.user.is_staff:
                        if not transmittal.destination_location or transmittal.destination_location != custodian_location:
                            continue
                    
                    # Update status to arrived
                    transmittal.status = 'arrived'
                    transmittal.arrived_at = timezone.now()
                    transmittal.arrived_by = request.user
                    transmittal.save()
                    
                    print(f"[DEBUG] Bulk mark arrived for transmittal {transmittal.pk}")
                    updated_count += 1
                    
                except Transmittal.DoesNotExist:
                    continue
            
            if updated_count == 0:
                return JsonResponse({
                    'success': False,
                    'error': 'No transmittals were marked as arrived. Please check your permissions.'
                })
            
            return JsonResponse({
                'success': True,
                'message': f'{updated_count} transmittal(s) marked as arrived successfully.',
                'updated_count': updated_count
            })
        except Exception as e:
            print(f"[ERROR] Exception in bulk_arrived: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method.'
    }, status=400)


# ============================================================
# LEGACY (compatibility)
# ============================================================
@login_required(login_url='accounts:login')
def toggle_resolved(request, pk):
    """Toggle resolved status (legacy compatibility)."""
    transmittal = get_object_or_404(Transmittal, pk=pk)
    
    if not (request.user.is_staff or 
            request.user == transmittal.sender or 
            request.user == transmittal.recipient_id):
        messages.error(request, "Permission denied.")
        return redirect('transmittals:detail', pk=pk)

    if request.method == 'POST':
        transmittal.is_resolved = not transmittal.is_resolved
        transmittal.save()
        status = "resolved" if transmittal.is_resolved else "unresolved"
        messages.success(request, f"Transmittal marked as {status}.")
        
    return redirect('transmittals:detail', pk=pk)


# ============================================================
# CUSTODIAN VIEWS
# ============================================================
def is_custodian(user):
    """Check if user is a custodian"""
    if not user.is_authenticated:
        return False
    try:
        return user.profile.role == 'custodian'
    except Profile.DoesNotExist:
        return False


@never_cache
@login_required(login_url='accounts:login')
def custodian_dashboard(request):
    """Custodian dashboard showing transmittals for their location"""
    # Check if user is custodian
    if not is_custodian(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('accounts:dashboard')
    
    # Get custodian's location
    try:
        profile = request.user.profile
        location = profile.assigned_location
        if not location:
            messages.error(request, "No location assigned to you.")
            return redirect('accounts:dashboard')
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found.")
        return redirect('accounts:dashboard')
    
    # Get transmittals for this location
    in_transit = Transmittal.objects.filter(
        destination_location=location,
        status='in_transit',
        recipient_deleted=False
    ).order_by('-sent_at')
    
    arrived = Transmittal.objects.filter(
        destination_location=location,
        status='arrived',
        recipient_deleted=False
    ).order_by('-sent_at')
    
    picked = Transmittal.objects.filter(
        destination_location=location,
        status='picked',
        recipient_deleted=False
    ).order_by('-sent_at')
    
    received = Transmittal.objects.filter(
        destination_location=location,
        status='received',
        recipient_deleted=False
    ).order_by('-sent_at')
    
    context = {
        'location': location,
        'in_transit': in_transit,
        'arrived': arrived,
        'picked': picked,
        'received': received,
        'in_transit_count': in_transit.count(),
        'arrived_count': arrived.count(),
        'picked_count': picked.count(),
        'received_count': received.count(),
    }
    return render(request, 'transmittals/custodian_dashboard.html', context)


@never_cache
@login_required(login_url='accounts:login')
def custodian_in_transit(request):
    """View all in transit transmittals for custodian's location"""
    if not is_custodian(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('accounts:dashboard')
    
    try:
        profile = request.user.profile
        location = profile.assigned_location
        if not location:
            messages.error(request, "No location assigned to you.")
            return redirect('accounts:dashboard')
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found.")
        return redirect('accounts:dashboard')
    
    transmittals = Transmittal.objects.filter(
        destination_location=location,
        status='in_transit',
        recipient_deleted=False
    ).order_by('-sent_at')
    
    context = {
        'transmittals': transmittals,
        'location': location,
        'status_filter': 'in_transit',
    }
    return render(request, 'transmittals/custodian_list.html', context)


@never_cache
@login_required(login_url='accounts:login')
def custodian_arrived(request):
    """View all arrived transmittals for custodian's location"""
    if not is_custodian(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('accounts:dashboard')
    
    try:
        profile = request.user.profile
        location = profile.assigned_location
        if not location:
            messages.error(request, "No location assigned to you.")
            return redirect('accounts:dashboard')
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found.")
        return redirect('accounts:dashboard')
    
    transmittals = Transmittal.objects.filter(
        destination_location=location,
        status='arrived',
        recipient_deleted=False
    ).order_by('-sent_at')
    
    context = {
        'transmittals': transmittals,
        'location': location,
        'status_filter': 'arrived',
    }
    return render(request, 'transmittals/custodian_list.html', context)


@never_cache
@login_required(login_url='accounts:login')
def custodian_picked(request):
    """View all picked transmittals for custodian's location"""
    if not is_custodian(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('accounts:dashboard')
    
    try:
        profile = request.user.profile
        location = profile.assigned_location
        if not location:
            messages.error(request, "No location assigned to you.")
            return redirect('accounts:dashboard')
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found.")
        return redirect('accounts:dashboard')
    
    transmittals = Transmittal.objects.filter(
        destination_location=location,
        status='picked',
        recipient_deleted=False
    ).order_by('-sent_at')
    
    context = {
        'transmittals': transmittals,
        'location': location,
        'status_filter': 'picked',
    }
    return render(request, 'transmittals/custodian_list.html', context)


def export_received_to_excel(transmittals, location):
    """Export received transmittals to Excel file"""
    
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Received Transmittals"
        
        # Define styles
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Add header
        headers = ['Transmittal Number', 'Sender', 'Recipient', 'Description', 'Sent Date', 'Arrived Date', 'Received Date', 'Remarks', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Add data - evaluate queryset
        transmittal_list = list(transmittals)
        print(f"DEBUG: Processing {len(transmittal_list)} transmittals")
        
        for row_idx, transmittal in enumerate(transmittal_list, 2):
            try:
                # Convert timestamps to local timezone
                # Since USE_TZ = False, work with naive datetimes directly
                sent_at_local = transmittal.sent_at
                arrived_at_local = transmittal.arrived_at
                received_at_local = transmittal.received_at
                
                # Get sender name safely
                sender_name = ''
                if transmittal.sender:
                    try:
                        sender_name = str(transmittal.sender.profile.get_full_name())
                    except (AttributeError, Profile.DoesNotExist):
                        sender_name = str(transmittal.sender.username) if transmittal.sender.username else ''
                
                ws.cell(row=row_idx, column=1).value = str(transmittal.reference_number) if transmittal.reference_number else ''
                ws.cell(row=row_idx, column=2).value = sender_name
                ws.cell(row=row_idx, column=3).value = str(transmittal.recipient_name) if transmittal.recipient_name else ''
                ws.cell(row=row_idx, column=4).value = str(transmittal.description) if transmittal.description else ''
                ws.cell(row=row_idx, column=5).value = sent_at_local.strftime('%Y-%m-%d %H:%M') if sent_at_local else ''
                ws.cell(row=row_idx, column=6).value = arrived_at_local.strftime('%Y-%m-%d %H:%M') if arrived_at_local else ''
                ws.cell(row=row_idx, column=7).value = received_at_local.strftime('%Y-%m-%d %H:%M') if received_at_local else ''
                ws.cell(row=row_idx, column=8).value = str(transmittal.remarks) if transmittal.remarks else ''
                ws.cell(row=row_idx, column=9).value = str(transmittal.status).upper() if transmittal.status else ''
                
                # Apply borders and alignment
                for col in range(1, 10):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border
                    cell.alignment = left_align
            except Exception as e:
                print(f"ERROR processing row {row_idx}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 12
        
        # Create response using BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Received_Transmittals_{location.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        print(f"DEBUG: File created: {filename}")
        return response
    except Exception as e:
        print(f"CRITICAL ERROR in export: {e}")
        import traceback
        traceback.print_exc()
        raise


@never_cache
@login_required(login_url='accounts:login')
def custodian_received(request):
    """View all received transmittals for custodian's location"""
    if not is_custodian(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('accounts:dashboard')
    
    try:
        profile = request.user.profile
        location = profile.assigned_location
        if not location:
            messages.error(request, "No location assigned to you.")
            return redirect('accounts:dashboard')
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found.")
        return redirect('accounts:dashboard')
    
    transmittals = Transmittal.objects.filter(
        destination_location=location,
        status='received',
        recipient_deleted=False
    ).order_by('-sent_at')
    
    # Handle date range filtering
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        try:
            from datetime import datetime as dt
            date_from_obj = dt.strptime(date_from, '%Y-%m-%d')
            transmittals = transmittals.filter(sent_at__gte=date_from_obj)
        except:
            pass
            
    if date_to:
        try:
            from datetime import datetime as dt
            date_to_obj = dt.strptime(date_to, '%Y-%m-%d')
            # Add one day to include the entire day
            from datetime import timedelta
            date_to_obj = date_to_obj + timedelta(days=1)
            transmittals = transmittals.filter(sent_at__lt=date_to_obj)
        except:
            pass
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        try:
            print(f"DEBUG: Export request received. Transmittals count: {transmittals.count()}")
            response = export_received_to_excel(transmittals, location)
            print(f"DEBUG: Export response created successfully")
            return response
        except Exception as e:
            print(f"ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error generating Excel file: {str(e)}")
            return redirect('transmittals:custodian_received')
    
    context = {
        'transmittals': transmittals,
        'location': location,
        'status_filter': 'received',
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'transmittals/custodian_list.html', context)


@never_cache
@login_required(login_url='accounts:login')
def custodian_export_received(request):
    """Dedicated page for exporting received transmittals to Excel"""
    if not is_custodian(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('accounts:dashboard')
    
    try:
        profile = request.user.profile
        location = profile.assigned_location
        if not location:
            messages.error(request, "No location assigned to you.")
            return redirect('accounts:dashboard')
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found.")
        return redirect('accounts:dashboard')
    
    # Get status filter
    status_filter = request.GET.get('status', 'received')  # Default to 'received'
    
    # Get all transmittals for the location
    if status_filter == 'all':
        transmittals = Transmittal.objects.filter(
            destination_location=location,
            recipient_deleted=False
        ).order_by('-sent_at')
    else:
        # Filter by specific status
        transmittals = Transmittal.objects.filter(
            destination_location=location,
            status=status_filter,
            recipient_deleted=False
        ).order_by('-sent_at')
    
    # Handle date range filtering
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        try:
            from datetime import datetime as dt
            date_from_obj = dt.strptime(date_from, '%Y-%m-%d')
            transmittals = transmittals.filter(sent_at__gte=date_from_obj)
        except:
            pass
            
    if date_to:
        try:
            from datetime import datetime as dt
            date_to_obj = dt.strptime(date_to, '%Y-%m-%d')
            from datetime import timedelta
            date_to_obj = date_to_obj + timedelta(days=1)
            transmittals = transmittals.filter(sent_at__lt=date_to_obj)
        except:
            pass
    
    # Handle Excel export download
    if request.GET.get('download') == 'excel':
        try:
            print(f"DEBUG: Export request received. Status: {status_filter}, Transmittals count: {transmittals.count()}")
            response = export_received_to_excel(transmittals, location)
            print(f"DEBUG: Export response created successfully")
            return response
        except Exception as e:
            print(f"ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error generating Excel file: {str(e)}")
    
    context = {
        'transmittals': transmittals,
        'location': location,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'transmittal_count': transmittals.count(),
    }
    return render(request, 'transmittals/custodian_export_received.html', context)


@never_cache
@login_required(login_url='accounts:login')
def custodian_outgoing(request):
    """View all transmittals from custodian's location (origin location)"""
    if not is_custodian(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('accounts:dashboard')
    
    # Get custodian's assigned location
    custodian_profile = request.user.profile
    custodian_location = custodian_profile.assigned_location
    
    # Get filter parameters
    filter_id = request.GET.get('filter_id', '')
    status_filter_param = request.GET.get('status', '')  # 'in_transit', 'arrived', 'received', or empty for ALL
    
    # Get all transmittals where origin_location matches custodian's assigned location
    # This includes transmittals sent by:
    # 1. The custodian themselves
    # 2. Regular users from their location
    transmittals_base = Transmittal.objects.filter(
        origin_location=custodian_location,
        sender_deleted=False
    )
    
    # Get filtered transmittals
    transmittals = transmittals_base
    
    # Apply destination filter if provided
    if filter_id:
        transmittals = transmittals.filter(destination_location_id=filter_id)
    
    # Apply status filter if provided
    if status_filter_param:
        transmittals = transmittals.filter(status=status_filter_param)
    
    # Order by sent_at in descending order
    transmittals = transmittals.order_by('-sent_at')
    
    # Get all unique destination locations for the filter dropdown from base transmittals
    from transmittals.models import Location
    
    all_destination_locations = Location.objects.filter(
        id__in=transmittals_base.values_list('destination_location_id', flat=True).distinct()
    ).order_by('name')
    
    context = {
        'transmittals': transmittals,
        'status_filter': 'outgoing',
        'filter_id': filter_id,
        'status_filter_param': status_filter_param,
        'all_destination_locations': all_destination_locations,
    }
    return render(request, 'transmittals/custodian_list.html', context)


@never_cache
@login_required(login_url='accounts:login')
def transmittal_flow_info(request):
    """
    Display transmittal flow information for both internal and external systems,
    including policies.
    """
    context = {
        # Internal flow steps
        'internal_steps': [
            {
                'number': 1,
                'status': 'In Transit',
                'icon': 'local_shipping',
                'description': 'Transmittal is being sent from the sender to the destination location.'
            },
            {
                'number': 2,
                'status': 'Arrived',
                'icon': 'inventory_2',
                'description': 'Transmittal has arrived at the destination location and is being held by the custodian.'
            },
            {
                'number': 3,
                'status': 'Picked Up',
                'icon': 'back_hand',
                'description': 'The intended recipient has picked up the transmittal from the custodian.'
            },
            {
                'number': 4,
                'status': 'Received',
                'icon': 'check_circle',
                'description': 'The recipient has confirmed receipt of the transmittal with digital signature.'
            }
        ],
        # External flow steps (timeline)
        'external_steps': [
            {
                'number': 1,
                'title': 'Creation',
                'description': 'Sender creates the transmittal, selecting Type and providing recipient details. "For Return" requires an Expected Return Date and a Hard Deadline.',
                'dot_color': '',
            },
            {
                'number': 2,
                'title': 'In-Transit',
                'description': 'The item is on its way. The system awaits confirmation of delivery.',
                'dot_color': '',
            },
            {
                'number': 3,
                'title': 'Delivery Confirmation',
                'description': 'The recipient or sender marks the item as "Received" by uploading a Proof of Delivery (e.g., photo of signed receipt). For "For Keep" types, this closes the transmittal.',
                'dot_color': '',
            },
            {
                'number': 4,
                'title': 'Open Status (For Return only)',
                'description': 'After delivery, "For Return" transmittals enter "Open" status. The system monitors the return deadline.',
                'dot_color': '#f9a825',
            },
            {
                'number': 5,
                'title': 'Resolution',
                'description': (
                    'The transmittal is resolved through one of the following actions:'
                    '<br>• <strong>Full Return:</strong> All items returned. (Closes transmittal)'
                    '<br>• <strong>Partial Return:</strong> Some items returned. (Remains Open)'
                    '<br>• <strong>Paid Sample:</strong> Receiver keeps item and pays. (Closes transmittal)'
                    '<br>• <strong>Convert to Keep:</strong> Decision to make transfer permanent. (Closes transmittal)'
                ),
                'dot_color': '#34a853',
            },
        ],
    }
    return render(request, 'transmittals/flow_info.html', context)


# ============================================================================
# TRANSMITTAL REPORT & EXPORT (USER & ADMIN)
# ============================================================================

@login_required(login_url='accounts:login')
@never_cache
def user_transmittal_report(request):
    """
    User transmittal report page with filtering and Excel export.
    Shows all transmittals where user is sender or recipient.
    """
    user = request.user
    
    # Base queryset: transmittals where user is sender or recipient
    sent_transmittals = Transmittal.objects.filter(
        sender=user,
        sender_deleted=False
    ).order_by('-sent_at')
    
    received_transmittals = Transmittal.objects.filter(
        recipient_id=user,
        recipient_deleted=False
    ).order_by('-sent_at')
    
    # Combine querysets
    from django.db.models import Q
    transmittals = Transmittal.objects.filter(
        Q(sender=user, sender_deleted=False) |
        Q(recipient_id=user, recipient_deleted=False)
    ).order_by('-sent_at')
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    direction_filter = request.GET.get('direction', '')  # 'sent' or 'received'
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Apply direction filter
    if direction_filter == 'sent':
        transmittals = transmittals.filter(sender=user, sender_deleted=False)
    elif direction_filter == 'received':
        transmittals = transmittals.filter(recipient_id=user, recipient_deleted=False)
    
    # Apply status filter
    if status_filter:
        transmittals = transmittals.filter(status=status_filter)
    
    # Apply date range filter
    if date_from:
        try:
            from datetime import datetime as dt
            date_from_obj = dt.strptime(date_from, '%Y-%m-%d')
            transmittals = transmittals.filter(sent_at__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            from datetime import datetime as dt, timedelta
            date_to_obj = dt.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            transmittals = transmittals.filter(sent_at__lt=date_to_obj)
        except:
            pass
    
    # Handle Excel export download
    if request.GET.get('download') == 'excel':
        try:
            response = export_transmittal_to_excel(transmittals, user)
            return response
        except Exception as e:
            print(f"ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error generating Excel file: {str(e)}")
    
    context = {
        'transmittals': transmittals,
        'transmittal_count': transmittals.count(),
        'status_filter': status_filter,
        'direction_filter': direction_filter,
        'date_from': date_from,
        'date_to': date_to,
        'page_title': 'Transmittal Reports',
    }
    return render(request, 'transmittals/user_report.html', context)


@login_required(login_url='accounts:login')
@never_cache
def custodian_transmittal_report(request):
    """
    Custodian transmittal report page with filtering and Excel export.
    Shows all transmittals for the custodian's assigned location.
    """
    if not is_custodian(request.user):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('accounts:dashboard')
    
    try:
        profile = request.user.profile
        location = profile.assigned_location
        if not location:
            messages.error(request, "No location assigned to you.")
            return redirect('accounts:dashboard')
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found.")
        return redirect('accounts:dashboard')
    
    # Get all transmittals for the location
    transmittals = Transmittal.objects.filter(
        destination_location=location,
        recipient_deleted=False
    ).order_by('-sent_at')
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Apply status filter
    if status_filter:
        transmittals = transmittals.filter(status=status_filter)
    
    # Apply date range filter
    if date_from:
        try:
            from datetime import datetime as dt
            date_from_obj = dt.strptime(date_from, '%Y-%m-%d')
            transmittals = transmittals.filter(sent_at__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            from datetime import datetime as dt, timedelta
            date_to_obj = dt.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            transmittals = transmittals.filter(sent_at__lt=date_to_obj)
        except:
            pass
    
    # Handle Excel export download
    if request.GET.get('download') == 'excel':
        try:
            response = export_received_to_excel(transmittals, location)
            return response
        except Exception as e:
            print(f"ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error generating Excel file: {str(e)}")
    
    context = {
        'transmittals': transmittals,
        'transmittal_count': transmittals.count(),
        'location': location,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'page_title': 'Custodian Transmittal Reports',
    }
    return render(request, 'transmittals/custodian_report.html', context)


def export_transmittal_to_excel(transmittals, user):
    """Export internal transmittals to Excel file."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transmittals"

        # Define styles
        header_fill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Define headers
        headers = [
            'Reference Number',
            'From (Sender)',
            'To (Recipient)',
            'Department',
            'Description',
            'Status',
            'Sent Date',
            'Received Date',
            'Location',
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Add data
        for row_idx, t in enumerate(transmittals, 2):
            try:
                ws.cell(row=row_idx, column=1).value = str(t.reference_number) if t.reference_number else ''
                ws.cell(row=row_idx, column=2).value = t.sender.get_full_name() or t.sender.username
                ws.cell(row=row_idx, column=3).value = t.recipient_id.get_full_name() or t.recipient_id.username if t.recipient_id else ''
                ws.cell(row=row_idx, column=4).value = t.sender.profile.department.name if hasattr(t.sender, 'profile') and t.sender.profile.department else ''
                ws.cell(row=row_idx, column=5).value = str(t.description) if t.description else ''
                ws.cell(row=row_idx, column=6).value = t.get_status_display().upper() if t.status else ''
                ws.cell(row=row_idx, column=7).value = t.sent_at.strftime('%Y-%m-%d %H:%M') if t.sent_at else ''
                ws.cell(row=row_idx, column=8).value = t.received_at.strftime('%Y-%m-%d %H:%M') if t.received_at else ''
                ws.cell(row=row_idx, column=9).value = t.destination_location.name if t.destination_location else ''

                # Apply borders and alignment
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border
                    cell.alignment = left_align
            except Exception as e:
                print(f"ERROR processing row {row_idx}: {e}")
                continue

        # Adjust column widths
        col_widths = [18, 20, 20, 20, 30, 14, 18, 18, 18]
        for i, width in enumerate(col_widths):
            ws.column_dimensions[chr(65 + i)].width = width

        # Create response
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_content = excel_buffer.getvalue()
        excel_buffer.seek(0)

        filename = f'Transmittals_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            excel_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
        response['Content-Length'] = len(excel_content)
        return response
    except Exception as e:
        print(f"CRITICAL ERROR in export: {e}")
        import traceback
        traceback.print_exc()
        raise

