"""
Script to create Excel import template
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create a template Excel file
wb = Workbook()
ws = wb.active
ws.title = "Transmittals"

# Define styles
header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Define columns
columns = [
    'Reference Number',
    'Sender Email',
    'Recipient Name',
    'Recipient Email',
    'Recipient Company',
    'Origin Location',
    'Destination Location',
    'Description',
    'Status',
]

# Write headers
for col_num, column_title in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = column_title
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Add sample row
sample_data = [
    'TRN-001',
    'admin@cdc.com',
    'John Doe',
    'john@cdc.com',
    'CDC',
    'Araneta',
    'Head Office',
    'Monthly Report',
    'in_transit',
]

for col_num, cell_value in enumerate(sample_data, 1):
    cell = ws.cell(row=2, column=col_num)
    cell.value = cell_value

# Set column widths
for col_num in range(1, len(columns) + 1):
    ws.column_dimensions[chr(64 + col_num)].width = 18

# Save the file
template_path = 'transmittal_import_template.xlsx'
wb.save(template_path)
print(f"✓ Template created: {template_path}")
