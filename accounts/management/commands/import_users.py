from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from django.contrib.auth.models import User
from accounts.models import Profile
from transmittals.models import Location
import re


def _make_unique_username(base):
    """Generate a unique username from a base string"""
    base = re.sub(r'\W+', '', base).lower() or 'user'
    username = base[:30]
    i = 1
    while User.objects.filter(username=username).exists():
        username = f"{base[:28]}{i}"
        i += 1
    return username


def _validate_email(email):
    """Validate email format"""
    if not email or not isinstance(email, str):
        return False, "Email is required"
    email = email.strip()
    if '@' not in email or '.' not in email.split('@')[1]:
        return False, "Invalid email format"
    if len(email) > 254:
        return False, "Email too long"
    return True, ""


def _validate_password(password):
    """Validate password requirements"""
    if not password or not isinstance(password, str):
        return False, "Password required"
    if len(password) < 4:
        return False, f"Password must be >=4 chars (got {len(password)})"
    return True, ""


def _validate_name(name, field_name):
    """Validate first/last name"""
    if not name or not isinstance(name, str):
        return False, f"{field_name} required"
    name = name.strip()
    if len(name) < 2:
        return False, f"{field_name} too short"
    if len(name) > 150:
        return False, f"{field_name} too long"
    return True, ""


def _validate_contact(contact):
    """Validate contact/phone number"""
    if not contact or not isinstance(contact, str):
        return False, "Contact required"
    contact = contact.strip()
    cleaned = re.sub(r'[\s\-\(\)\+]', '', contact)
    if not cleaned:
        return False, "Contact cannot be empty/formatting only"
    if len(contact) > 20:
        return False, "Contact too long"
    return True, ""


def _validate_text_field(value, field_name, max_length=100):
    """Generic text field validation"""
    if not value or not isinstance(value, str):
        return False, f"{field_name} required"
    value = value.strip()
    if not value:
        return False, f"{field_name} cannot be empty"
    if len(value) > max_length:
        return False, f"{field_name} too long"
    return True, ""


def _find_location(assigned_name):
    """Find location by name"""
    if not assigned_name or not isinstance(assigned_name, str):
        return None
    assigned_name = assigned_name.strip()
    location = Location.objects.filter(name__iexact=assigned_name, is_active=True).first()
    if location:
        return location
    location = Location.objects.filter(name__icontains=assigned_name, is_active=True).first()
    return location


class Command(BaseCommand):
    help = "Import users from Excel with comprehensive validation"

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help='Path to Excel file')
        parser.add_argument('--sheet', type=str, default=None)
        parser.add_argument('--dry-run', action='store_true', help='Validate without writing')
        parser.add_argument('--limit', type=int, default=None, help='Limit number of rows to import')

    def handle(self, *args, **options):
        path = options['file']
        sheet = options['sheet']
        dry_run = options['dry_run']
        limit = options.get('limit')

        try:
            wb = load_workbook(path, data_only=True)
            ws = wb[sheet] if sheet else wb.active
        except Exception as e:
            self.stderr.write(f"Error loading file: {e}")
            return

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.stderr.write("Excel file is empty")
            return
        
        if limit:
            rows = rows[:limit+1]  # +1 to include header

        header = [str(h).strip() if h is not None else '' for h in rows[0]]
        col_index = {name: idx for idx, name in enumerate(header)}

        required = ['Username', 'Email Address', 'Password', 'First Name', 'Last Name', 'Contact Number', 'Company Name', 'Department', 'Assigned Location']
        missing = [c for c in required if c not in col_index]
        if missing:
            self.stderr.write(f"Missing columns: {', '.join(missing)}")
            return

        self.stdout.write(f"\n{'='*70}")
        self.stdout.write(f"File: {path}")
        self.stdout.write(f"Mode: {'DRY-RUN' if dry_run else 'LIVE'}")
        self.stdout.write(f"{'='*70}\n")

        created = 0
        skipped = 0
        errors = 0

        for row_num, row in enumerate(rows[1:], start=2):
            row_errors = []
            
            username = str(row[col_index.get('Username', 0)] or '').strip()
            email = str(row[col_index.get('Email Address', 0)] or '').strip()
            password = str(row[col_index.get('Password', 0)] or '').strip()
            first_name = str(row[col_index.get('First Name', 0)] or '').strip()
            last_name = str(row[col_index.get('Last Name', 0)] or '').strip()
            contact = str(row[col_index.get('Contact Number', 0)] or '').strip()
            company = str(row[col_index.get('Company Name', 0)] or '').strip()
            department = str(row[col_index.get('Department', 0)] or '').strip()
            assigned_name = str(row[col_index.get('Assigned Location', 0)] or '').strip()

            if not any([email, password, first_name, last_name]):
                skipped += 1
                continue

            # Validate username
            valid, msg = _validate_text_field(username, "Username", 150)
            if not valid:
                row_errors.append(f"Username: {msg}")
            valid, msg = _validate_email(email)
            if not valid:
                row_errors.append(f"Email: {msg}")
            
            valid, msg = _validate_password(password)
            if not valid:
                row_errors.append(f"Password: {msg}")
            
            valid, msg = _validate_name(first_name, "FirstName")
            if not valid:
                row_errors.append(f"FirstName: {msg}")
            
            valid, msg = _validate_name(last_name, "LastName")
            if not valid:
                row_errors.append(f"LastName: {msg}")
            
            valid, msg = _validate_contact(contact)
            if not valid:
                row_errors.append(f"Contact: {msg}")
            
            valid, msg = _validate_text_field(company, "Company", 100)
            if not valid:
                row_errors.append(f"Company: {msg}")
            
            valid, msg = _validate_text_field(department, "Department", 100)
            if not valid:
                row_errors.append(f"Department: {msg}")
            
            location = _find_location(assigned_name)
            if not location:
                row_errors.append(f"Location: '{assigned_name}' not found/inactive")
            
            if User.objects.filter(email=email).exists():
                row_errors.append("Email: already exists")
            
            if User.objects.filter(username=username).exists():
                row_errors.append("Username: already exists")
            
            if row_errors:
                errors += 1
                self.stderr.write(f"[ERROR] Row {row_num} ({email}): {'; '.join(row_errors)}")
                continue

            try:
                if dry_run:
                    self.stdout.write(f"[OK] Row {row_num}: {email} -> {username} (would create)")
                else:
                    user = User.objects.create_user(username=username, email=email, password=password, first_name=first_name, last_name=last_name)
                    user.is_active = True
                    user.save()
                    
                    profile, created_profile = Profile.objects.get_or_create(
                        user=user,
                        defaults={
                            'contact': contact,
                            'company': company,
                            'department': department,
                            'location': assigned_name,
                            'address': assigned_name,
                            'assigned_location': location,
                            'role': 'user',
                            'status': 'pending'
                        }
                    )
                    
                    # Prevent email notification by setting original status
                    profile._original_status = 'pending'
                    
                    if not created_profile:
                        # Update existing profile
                        profile._original_status = profile.status  # Preserve current status to prevent email
                    profile.contact = contact
                    profile.company = company
                    profile.department = department
                    profile.location = assigned_name
                    profile.address = assigned_name
                    profile.assigned_location = location
                    profile.status = 'pending'
                    profile.save(update_fields=['contact', 'company', 'department', 'location', 'address', 'assigned_location', 'status'])
                    created += 1
                    self.stdout.write(f"[OK] Row {row_num} | Email: {email} | Username: {username}")
            
            except Exception as exc:
                errors += 1
                self.stderr.write(f"[ERROR] Row {row_num}: Exception - {str(exc)}")

        self.stdout.write(f"\n{'='*70}")
        self.stdout.write(f"Created: {created} | Skipped: {skipped} | Errors: {errors}")
        self.stdout.write(f"Status: All users imported with 'pending' status")
        self.stdout.write(f"{'='*70}\n")
