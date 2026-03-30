"""
Admin Configuration for User Registration System

This module configures the Django admin interface for managing
users and their profiles.
"""

from django import forms
from django.contrib import admin
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.urls import path
from django.template.response import TemplateResponse
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib import messages
from .models import Profile, UserActivity, ActiveSession, Department
from .email_utils import send_approval_email, send_rejection_email
from django.utils.html import format_html
from openpyxl import load_workbook, Workbook
import io
from datetime import datetime


class ProfileAdminForm(forms.ModelForm):
    """Custom form for ProfileAdmin to display department as dropdown"""
    department = forms.ModelChoiceField(
        queryset=Department.objects.filter(is_active=True),
        required=False,
        label='Department',
        empty_label='-- Select Department --'
    )
    
    class Meta:
        model = Profile
        fields = '__all__'
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Set initial department value from the instance
        if self.instance.pk and self.instance.department:
            self.fields['department'].initial = self.instance.department


@admin.register(Department)
class DepartmentAdmin(admin.ModelAdmin):
    """
    Admin interface for Department management
    """
    list_display = ('name', 'is_active', 'created_at')
    list_filter = ('is_active', 'created_at')
    search_fields = ('name',)
    ordering = ('name',)


@admin.register(Profile)
class ProfileAdmin(admin.ModelAdmin):
    """
    Custom Admin Interface for User Profiles
    
    Features:
    - Display profile information in list view
    - Filter by registration status
    - Search by username, email, or name
    - Quick approve/reject actions
    - Inline editing of profile information
    - View registration details
    """
    
    form = ProfileAdminForm
    
    # Fields to display in list view
    list_display = [
        'username',
        'full_name',
        'email',
        'role_badge',
        'company',
        'signature_preview',
        'status_badge',
        'created_at',
        'actions_column'
    ]
    
    # Fields that can be searched
    search_fields = [
        'user__username',
        'user__email',
        'user__first_name',
        'user__last_name',
        'company',
        'department__name'
    ]
    
    # Filter options in sidebar
    list_filter = [
        'status',
        'role',
        'created_at',
        'department',
        'company'
    ]
    
    # Read-only fields
    readonly_fields = [
        'user',
        'created_at',
        'updated_at',
        'get_user_details'
    ]
    
    # Fields to display in edit form
    fieldsets = (
        ('User Information', {
            'fields': ('user', 'get_user_details')
        }),
        ('Contact Information', {
            'fields': ('contact',)
        }),
        ('Organization', {
            'fields': ('company', 'department', 'location')
        }),
        ('Address', {
            'fields': ('address',)
        }),
        ('Role Assignment', {
            'fields': ('role', 'assigned_location'),
            'description': 'Assign user role and location for custodians'
        }),
        ('Digital Signature', {
            'fields': ('digital_signature',),
            'description': 'Upload or view user digital signature (PNG, JPG, GIF)'
        }),
        ('Registration Status', {
            'fields': ('status', 'admin_notes')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    # Ordering in list view
    ordering = ['-created_at']
    
    # Enable pagination
    list_per_page = 25
    
    # Custom actions
    actions = ['approve_selected', 'reject_selected']
    
    def username(self, obj):
        """Display username from related User"""
        return obj.user.username
    username.short_description = 'Username'
    
    def full_name(self, obj):
        """Display user's full name"""
        return obj.get_full_name()
    full_name.short_description = 'Full Name'
    
    def role_badge(self, obj):
        """Display colored role badge"""
        colors = {
            'user': '#6C757D',       # Gray
            'custodian': '#007BFF',  # Blue
            'admin': '#DC3545'       # Red
        }
        color = colors.get(obj.role, '#6C757D')
        
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-weight: bold; font-size: 11px;">{}</span>',
            color,
            obj.get_role_display()
        )
    role_badge.short_description = 'Role'
    
    def email(self, obj):
        """Display user's email"""
        return obj.user.email
    email.short_description = 'Email'
    
    def status_badge(self, obj):
        """Display colored status badge"""
        colors = {
            'pending': '#FFC107',    # Yellow/Orange
            'approved': '#28A745',   # Green
            'rejected': '#DC3545'    # Red
        }
        color = colors.get(obj.status, '#6C757D')  # Default gray
        
        return format_html(
            '<span style="background-color: {}; color: white; padding: 5px 10px; '
            'border-radius: 3px; font-weight: bold;">{}</span>',
            color,
            obj.get_status_display()
        )
    status_badge.short_description = 'Status'
    
    def signature_preview(self, obj):
        """Display signature image preview"""
        if obj.digital_signature:
            signature_url = obj.digital_signature.url
            return format_html(
                '<img src="{}" style="max-width: 100px; max-height: 50px; border: 1px solid #ddd; border-radius: 3px; padding: 2px;" title="Signature of {}" />',
                signature_url,
                obj.get_full_name()
            )
        return format_html(
            '<span style="color: #999; font-style: italic;">No signature</span>'
        )
    signature_preview.short_description = 'Digital Signature'
    
    def actions_column(self, obj):
        """Display quick action buttons"""
        if obj.status == 'pending':
            return format_html(
                '<a class="button" style="background-color: #28A745; margin-right: 5px;" '
                'href="/admin/accounts/profile/{}/approve/">Approve</a>'
                '<a class="button" style="background-color: #DC3545;" '
                'href="/admin/accounts/profile/{}/reject/">Reject</a>',
                obj.id, obj.id
            )
        return '-'
    actions_column.short_description = 'Quick Actions'
    
    def get_user_details(self, obj):
        """Display detailed user information"""
        user = obj.user
        return format_html(
            '<b>Username:</b> {}<br>'
            '<b>Email:</b> {}<br>'
            '<b>First Name:</b> {}<br>'
            '<b>Last Name:</b> {}<br>'
            '<b>Active:</b> {}<br>'
            '<b>Staff:</b> {}',
            user.username,
            user.email,
            user.first_name,
            user.last_name,
            'Yes' if user.is_active else 'No',
            'Yes' if user.is_staff else 'No'
        )
    get_user_details.short_description = 'User Account Details'
    
    @admin.action(description='Approve selected users')
    def approve_selected(self, request, queryset):
        """Bulk action to approve selected users"""
        for profile in queryset:
            if profile.status != 'approved':
                profile.status = 'approved'
                profile.save()
                send_approval_email(profile)
        
        count = queryset.count()
        self.message_user(request, f'{count} user(s) approved successfully!')
    
    @admin.action(description='Reject selected users')
    def reject_selected(self, request, queryset):
        """Bulk action to reject selected users"""
        for profile in queryset:
            if profile.status != 'rejected':
                profile.status = 'rejected'
                profile.save()
                send_rejection_email(profile)
        
        count = queryset.count()
        self.message_user(request, f'{count} user(s) rejected!')
    
    def save_model(self, request, obj, form, change):
        """
        Override save_model to provide feedback when status changes trigger emails.
        """
        # Check if status field changed
        if change and 'status' in form.changed_data:
            old_status = form.initial.get('status')
            new_status = obj.status
            
            # Save the model first
            super().save_model(request, obj, form, change)
            
            # Provide feedback based on status change
            if old_status == 'pending' and new_status == 'approved':
                self.message_user(
                    request, 
                    f'User {obj.user.username} has been approved! An approval email has been sent to {obj.user.email}.',
                    level='SUCCESS'
                )
            elif old_status == 'pending' and new_status == 'rejected':
                self.message_user(
                    request, 
                    f'User {obj.user.username} has been rejected. A rejection email has been sent to {obj.user.email}.',
                    level='WARNING'
                )
        else:
            # No status change, save normally
            super().save_model(request, obj, form, change)


# Extend the User admin to show profile information
class ProfileInline(admin.StackedInline):
    """Inline admin for viewing/editing profile alongside User"""
    model = Profile
    fields = ('contact', 'company', 'department', 'address', 'role', 'assigned_location', 'status', 'admin_notes')
    extra = 0


class CustomUserAdmin(BaseUserAdmin):
    """Extended User Admin with Profile information"""
    inlines = [ProfileInline]
    
    list_display = [
        'username',
        'email',
        'first_name',
        'last_name',
        'get_assigned_location',
        'get_department',
        'is_staff',
        'is_active',
        'get_profile_status'
    ]
    
    list_filter = [
        'is_staff',
        'is_active',
        'profile__status',
        'profile__assigned_location',
        'profile__department',
        'date_joined'
    ]
    
    search_fields = [
        'username',
        'email',
        'first_name',
        'last_name',
        'profile__assigned_location__name',
        'profile__department__name',
    ]
    
    actions = ['approve_users', 'reject_users', 'export_users_to_excel']
    
    def get_urls(self):
        """Add custom URLs for user import"""
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_users_excel_view), 
                 name='user_import_excel'),
        ]
        return custom_urls + urls
    
    def import_users_excel_view(self, request):
        """Handle Excel file upload and import users"""
        if request.method == 'POST':
            if 'excel_file' not in request.FILES:
                messages.error(request, "❌ No file selected. Please choose an Excel file.")
                return HttpResponseRedirect(request.path)
            
            excel_file = request.FILES['excel_file']
            
            # Validate file extension
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "❌ Invalid file format. Only .xlsx and .xls files are supported.")
                return HttpResponseRedirect(request.path)
            
            # Validate file size (max 5MB)
            if excel_file.size > 5 * 1024 * 1024:
                messages.error(request, "❌ File size cannot exceed 5MB.")
                return HttpResponseRedirect(request.path)
            
            try:
                # Read Excel file
                excel_data = io.BytesIO(excel_file.read())
                workbook = load_workbook(excel_data)
                worksheet = workbook.active
                
                imported = 0
                errors = []
                warnings = []
                
                # Get all rows
                rows = list(worksheet.iter_rows(min_row=2, values_only=True))
                
                if not rows or all(cell is None for row in rows for cell in row):
                    messages.error(request, "❌ Excel file is empty. Please add data rows.")
                    return HttpResponseRedirect(request.path)
                
                # Process each row
                for row_num, row in enumerate(rows, start=2):
                    try:
                        # Skip empty rows
                        if all(cell is None for cell in row):
                            continue
                        
                        # Map columns
                        username = row[0] if len(row) > 0 else None
                        email = row[1] if len(row) > 1 else None
                        first_name = row[2] if len(row) > 2 else None
                        last_name = row[3] if len(row) > 3 else None
                        password = row[4] if len(row) > 4 else None
                        contact = row[5] if len(row) > 5 else None
                        company = row[6] if len(row) > 6 else None
                        department_name = row[7] if len(row) > 7 else None
                        assigned_location_name = row[8] if len(row) > 8 else None
                        address = row[9] if len(row) > 9 else None
                        role = row[10] if len(row) > 10 else None
                        status = row[11] if len(row) > 11 else None
                        
                        # Validation: Username (Required)
                        if not username or str(username).strip() == '':
                            errors.append(f"Row {row_num}: ❌ Username is required")
                            continue
                        
                        username = str(username).strip()
                        
                        # Check if username already exists
                        if User.objects.filter(username=username).exists():
                            errors.append(f"Row {row_num}: ❌ Username '{username}' already exists")
                            continue
                        
                        # Validation: Email (Required)
                        if not email or str(email).strip() == '':
                            errors.append(f"Row {row_num}: ❌ Email is required")
                            continue
                        
                        email = str(email).strip()
                        
                        # Check if email already exists
                        if User.objects.filter(email=email).exists():
                            errors.append(f"Row {row_num}: ❌ Email '{email}' already exists")
                            continue
                        
                        # Validation: Password (Required)
                        if not password or str(password).strip() == '':
                            errors.append(f"Row {row_num}: ❌ Password is required")
                            continue
                        
                        password = str(password).strip()
                        
                        if len(password) < 8:
                            errors.append(f"Row {row_num}: ❌ Password must be at least 8 characters")
                            continue
                        
                        # Validation: First Name & Last Name (Optional)
                        first_name = str(first_name).strip() if first_name else ''
                        last_name = str(last_name).strip() if last_name else ''
                        
                        # Validation: Contact (Optional but recommended)
                        contact = str(contact).strip() if contact else None
                        if contact and not any(char.isdigit() for char in contact):
                            warnings.append(f"Row {row_num}: ⚠️ Contact '{contact}' doesn't contain digits")
                            contact = None
                        
                        # Validation: Company (Optional)
                        company = str(company).strip() if company else None
                        
                        # Get Department (Optional)
                        department = None
                        if department_name and str(department_name).strip():
                            department_name = str(department_name).strip()
                            try:
                                department = Department.objects.get(
                                    name__iexact=department_name,
                                    is_active=True
                                )
                            except Department.DoesNotExist:
                                available_depts = ', '.join([d.name for d in Department.objects.filter(is_active=True)[:5]])
                                warnings.append(f"Row {row_num}: ⚠️ Department '{department_name}' not found. Using None. Available: {available_depts}...")
                                department = None
                        
                        # Get Assigned Location (Optional)
                        assigned_location = None
                        if assigned_location_name and str(assigned_location_name).strip():
                            assigned_location_name = str(assigned_location_name).strip()
                            try:
                                from transmittals.models import Location
                                assigned_location = Location.objects.get(name__iexact=assigned_location_name)
                            except:
                                warnings.append(f"Row {row_num}: ⚠️ Location '{assigned_location_name}' not found. Using None.")
                                assigned_location = None
                        
                        # Address (Optional)
                        address = str(address).strip() if address else None
                        
                        # Role validation (Optional, default 'user')
                        valid_roles = ['user', 'custodian', 'admin']
                        if role and str(role).strip().lower() not in valid_roles:
                            warnings.append(f"Row {row_num}: ⚠️ Role '{role}' is invalid. Using 'user'. Valid options: {', '.join(valid_roles)}")
                            role = 'user'
                        else:
                            role = str(role).strip().lower() if role else 'user'
                        
                        # Status validation (Optional, default 'pending')
                        valid_statuses = ['pending', 'approved', 'rejected']
                        if status and str(status).strip().lower() not in valid_statuses:
                            warnings.append(f"Row {row_num}: ⚠️ Status '{status}' is invalid. Using 'pending'. Valid options: {', '.join(valid_statuses)}")
                            status = 'pending'
                        else:
                            status = str(status).strip().lower() if status else 'pending'
                        
                        # Create User
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            first_name=first_name,
                            last_name=last_name
                        )
                        
                        # Create Profile
                        profile = Profile.objects.create(
                            user=user,
                            contact=contact,
                            company=company,
                            department=department,
                            assigned_location=assigned_location,
                            address=address,
                            role=role,
                            status=status
                        )
                        
                        imported += 1
                        
                    except Exception as e:
                        errors.append(f"Row {row_num}: ❌ Unexpected error: {str(e)}")
                
                # Show results
                if imported > 0:
                    messages.success(
                        request, 
                        f"✓ Successfully imported {imported} user(s)"
                    )
                
                if warnings:
                    for warning in warnings[:10]:
                        messages.warning(request, warning)
                    if len(warnings) > 10:
                        messages.warning(request, f"... and {len(warnings) - 10} more warnings")
                
                if errors:
                    for error in errors[:10]:
                        messages.error(request, error)
                    if len(errors) > 10:
                        messages.error(request, f"... and {len(errors) - 10} more errors")
                
                return HttpResponseRedirect(request.path)
            
            except Exception as e:
                messages.error(request, f"❌ Error reading Excel file: {str(e)}")
                return HttpResponseRedirect(request.path)
        
        # Display import form
        from transmittals.models import Location
        context = {
            'title': 'Import Users from Excel',
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
            'has_add_permission': self.has_add_permission(request),
            'available_departments': Department.objects.filter(is_active=True),
            'available_locations': Location.objects.all(),
        }
        return TemplateResponse(request, 'admin/accounts/import_users_excel.html', context)
    
    def get_assigned_location(self, obj):
        """Display assigned location from profile"""
        try:
            location = obj.profile.assigned_location
            if location:
                return location.name
            return '-'
        except Profile.DoesNotExist:
            return '-'
    get_assigned_location.short_description = 'Assigned Location'
    
    def get_department(self, obj):
        """Display department from profile"""
        try:
            department = obj.profile.department
            if department:
                return department.name
            return '-'
        except Profile.DoesNotExist:
            return '-'
    get_department.short_description = 'Department'
    get_department.admin_order_field = 'profile__department'
    
    def get_profile_status(self, obj):
        """Display profile status in User list"""
        try:
            status = obj.profile.status
            colors = {
                'pending': '#FFC107',
                'approved': '#28A745',
                'rejected': '#DC3545'
            }
            color = colors.get(status, '#6C757D')
            return format_html(
                '<span style="background-color: {}; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-size: 11px;">{}</span>',
                color,
                status.upper()
            )
        except Profile.DoesNotExist:
            return '-'
    get_profile_status.short_description = 'Registration Status'
    
    def get_search_results(self, request, queryset, search_term):
        """
        Override to improve search/filter combination.
        Ensures search works correctly with location filter.
        """
        from django.db.models import Q
        
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        
        # If there's a search term and location filter is applied, ensure proper matching
        if search_term:
            q_objects = Q()
            for search_field in self.search_fields:
                if search_field == 'profile__assigned_location__name':
                    # Case-insensitive search for location name
                    q_objects |= Q(profile__assigned_location__name__icontains=search_term)
                else:
                    # Case-insensitive search for other fields
                    field_lookup = f'{search_field}__icontains'
                    q_objects |= Q(**{field_lookup: search_term})
            
            queryset = queryset.filter(q_objects).distinct()
        
        return queryset, True

    @admin.action(description='📥 Export selected users to Excel')
    def export_users_to_excel(self, request, queryset):
        """Export selected users to Excel file"""
        # Create workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Users"
        
        # Create headers
        headers = [
            'Username',
            'Email',
            'First Name',
            'Last Name',
            'Contact',
            'Company',
            'Department',
            'Assigned Location',
            'Address',
            'Role',
            'Status',
            'Date Joined'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = cell.font.copy()
            cell.font = cell.font.copy()
            from openpyxl.styles import Font, PatternFill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write data
        for row_num, user in enumerate(queryset, 2):
            try:
                profile = user.profile
            except Profile.DoesNotExist:
                profile = None
            
            row_data = [
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                profile.contact if profile else '',
                profile.company if profile else '',
                profile.department.name if profile and profile.department else '',
                profile.assigned_location.name if profile and profile.assigned_location else '',
                profile.address if profile else '',
                profile.role if profile else 'user',
                profile.status if profile else 'pending',
                user.date_joined.strftime("%Y-%m-%d %H:%M") if user.date_joined else '',
            ]
            
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num).value = value
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response['Content-Disposition'] = f'attachment; filename="Users_{timestamp}.xlsx"'
        
        workbook.save(response)
        return response

    @admin.action(description='Approve selected users')
    def approve_users(self, request, queryset):
        """Bulk action to approve selected users from User admin"""
        count = 0
        for user in queryset:
            try:
                profile = user.profile
                if profile.status != 'approved':
                    profile.status = 'approved'
                    profile.save()
                    send_approval_email(profile)
                    count += 1
            except Profile.DoesNotExist:
                pass
        
        self.message_user(request, f'{count} user(s) approved successfully!')
    
    @admin.action(description='Reject selected users')
    def reject_users(self, request, queryset):
        """Bulk action to reject selected users from User admin"""
        count = 0
        for user in queryset:
            try:
                profile = user.profile
                if profile.status != 'rejected':
                    profile.status = 'rejected'
                    profile.save()
                    send_rejection_email(profile)
                    count += 1
            except Profile.DoesNotExist:
                pass
        
        self.message_user(request, f'{count} user(s) rejected!')


# Unregister and Register User
try:
    admin.site.unregister(User)
except admin.sites.NotRegistered:
    pass

admin.site.register(User, CustomUserAdmin)


# Customize admin site
admin.site.site_header = "Email System Administration"
admin.site.site_title = "Email System Admin"
admin.site.index_title = "Welcome to Email System Admin Panel"


# =====================
# User Activity Tracking Admin
# =====================

@admin.register(UserActivity)
class UserActivityAdmin(admin.ModelAdmin):
    """
    Admin interface for viewing user activity logs
    Shows only login and logout activities
    """
    list_display = ['user', 'activity_type_badge', 'timestamp', 'ip_address']
    list_filter = ['timestamp', 'user']
    search_fields = ['user__username', 'user__email', 'ip_address']
    readonly_fields = ['user', 'activity_type', 'page_url', 'description', 'ip_address', 'user_agent', 'timestamp']
    date_hierarchy = 'timestamp'
    
    def get_queryset(self, request):
        """Filter queryset to show only login and logout activities"""
        queryset = super().get_queryset(request)
        return queryset.filter(activity_type__in=['login', 'logout']).order_by('-timestamp')
    
    def activity_type_badge(self, obj):
        """Display activity type with color coding"""
        colors = {
            'login': '#28a745',
            'logout': '#dc3545',
        }
        color = colors.get(obj.activity_type, '#6c757d')
        badge_text = '📍 Login' if obj.activity_type == 'login' else '🚪 Logout'
        return format_html(
            '<span style="background-color: {}; color: white; padding: 5px 12px; border-radius: 4px; font-weight: 600;">{}</span>',
            color,
            badge_text
        )
    activity_type_badge.short_description = 'Activity Type'
    
    def has_add_permission(self, request):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser


@admin.register(ActiveSession)
class ActiveSessionAdmin(admin.ModelAdmin):
    """
    Admin interface for viewing active user sessions
    Only shows users who are truly online (within session timeout window)
    """
    list_display = ['user_full_name', 'login_time', 'last_activity_display', 'session_duration', 'ip_address', 'is_online_badge']
    list_filter = ['login_time']
    search_fields = ['user__username', 'user__email', 'ip_address']
    readonly_fields = ['user', 'login_time', 'last_activity', 'ip_address', 'user_agent', 'session_key', 'get_session_duration']
    
    def get_queryset(self, request):
        """Only show sessions marked as active"""
        queryset = super().get_queryset(request)
        return queryset.filter(is_active=True).order_by('-last_activity')
    
    def user_full_name(self, obj):
        """Display user full name"""
        return f"{obj.user.get_full_name()} ({obj.user.username})"
    user_full_name.short_description = 'User'
    
    def last_activity_display(self, obj):
        """Display last activity time"""
        return obj.last_activity.strftime("%Y-%m-%d %H:%M:%S")
    last_activity_display.short_description = 'Last Activity'
    
    def is_online_badge(self, obj):
        """Display online status badge based on real-time activity check"""
        if obj.is_truly_online():
            return format_html(
                '<span style="background-color: #28a745; color: white; padding: 5px 12px; border-radius: 4px; font-weight: 600;">🟢 Online</span>'
            )
        return format_html(
            '<span style="background-color: #dc3545; color: white; padding: 5px 12px; border-radius: 4px; font-weight: 600;">🔴 Offline</span>'
        )
    is_online_badge.short_description = 'Status'
    
    def session_duration(self, obj):
        """Display session duration"""
        return obj.get_session_duration()
    session_duration.short_description = 'Session Duration'
    
    def has_add_permission(self, request):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser
